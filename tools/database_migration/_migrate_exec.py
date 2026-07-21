"""Copia ejecutable temporal de migrate_json_to_excel.py.

Existe solo como workaround de un problema de caché de sincronización
de archivos en la sesión de trabajo. El archivo canónico del proyecto
es ``migrate_json_to_excel.py`` (contenido idéntico). Puede borrarse.
"""
from __future__ import annotations

import json
from pathlib import Path

import bes

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
# tools/database_migration/ -> raiz del repo
_REPO_ROOT = BASE_DIR.parent.parent
# Los catalogos JSON son datos del paquete instalable: se resuelven a
# traves de el, no por ruta relativa al repo.
_BES_PKG = Path(bes.__file__).parent
CATALOG_DIR = _BES_PKG / "catalogs"
OUTPUT_DIR = BASE_DIR / "data_excel"

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
CELL_FONT = Font(name=FONT)


def make_id(entry: dict) -> str:
    parts = [entry["manufacturer"], entry.get("series", ""), entry["model"]]
    return "_".join(str(p) for p in parts if p)


def write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = CELL_FONT
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(width + 3, 60)
    ws.freeze_panes = "A2"


def write_readme(wb, title: str, descriptions: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("README")
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A3"] = "Columna / Hoja"
    ws["B3"] = "Descripción y unidades"
    for cell in (ws["A3"], ws["B3"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, (col, desc) in enumerate(descriptions, start=4):
        ws[f"A{i}"] = col
        ws[f"B{i}"] = desc
        ws[f"A{i}"].font = Font(name=FONT, bold=True)
        ws[f"B{i}"].font = CELL_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100


def load_json(name: str, key: str) -> list[dict]:
    with open(CATALOG_DIR / name, encoding="utf-8") as f:
        return json.load(f)[key]


def migrate_pumps() -> None:
    pumps = load_json("pumps.json", "pumps")
    wb = Workbook()
    ws = wb.active
    ws.title = "pumps"
    headers = ["pump_id", "manufacturer", "series", "model", "od_inches",
               "min_flow_bpd", "max_flow_bpd", "bep_flow_bpd", "max_stages",
               "housing_options", "source"]
    rows = [[make_id(p), p["manufacturer"], p["series"], p["model"],
             p["od_inches"], p["min_flow_bpd"], p["max_flow_bpd"],
             p["bep_flow_bpd"], p["max_stages"],
             ", ".join(str(h) for h in p["housing_options"]),
             p.get("_source", "")] for p in pumps]
    write_sheet(ws, headers, rows)

    ws2 = wb.create_sheet("performance_curves")
    headers2 = ["pump_id", "flow_bpd", "head_ft_per_stage", "hp_per_stage",
                "efficiency"]
    rows2 = [[make_id(p), pt["flow_bpd"], pt["head_ft_per_stage"],
              pt["hp_per_stage"], pt["efficiency"]]
             for p in pumps for pt in p["performance_curve"]]
    write_sheet(ws2, headers2, rows2)

    write_readme(wb, "Catálogo de bombas electrosumergibles (BES/ESP)", [
        ("Hoja 'pumps'", "Una fila por modelo de bomba (datos generales)."),
        ("Hoja 'performance_curves'", "Curva de rendimiento: una fila por punto (caudal, altura, potencia, eficiencia). Se liga a la bomba por pump_id."),
        ("pump_id", "Clave única legible: fabricante_serie_modelo. NO editar sin actualizar la hoja de curvas."),
        ("manufacturer / series / model", "Fabricante, serie (diámetro nominal) y modelo."),
        ("od_inches", "Diámetro externo de la bomba [pulgadas]. Debe ser menor al ID del casing."),
        ("min_flow_bpd / max_flow_bpd", "Rango operativo recomendado [barriles/día]."),
        ("bep_flow_bpd", "Caudal de máxima eficiencia (Best Efficiency Point) [bpd]."),
        ("max_stages", "Número máximo de etapas por consideraciones de housing."),
        ("housing_options", "Cantidades de etapas disponibles por housing, separadas por comas."),
        ("flow_bpd", "Caudal del punto de curva [bpd]."),
        ("head_ft_per_stage", "Altura dinámica entregada por etapa [pies/etapa]."),
        ("hp_per_stage", "Potencia al freno requerida por etapa [HP/etapa] (fluido gravedad específica 1.0)."),
        ("efficiency", "Eficiencia hidráulica de la bomba [fracción 0-1]."),
        ("source", "Referencia de origen del dato (trazabilidad: Brown Vol. 2B, catálogo del fabricante, etc.)."),
    ])
    wb.save(OUTPUT_DIR / "pumps.xlsx")


def migrate_cables() -> None:
    cables = load_json("cables.json", "cables")
    wb = Workbook()
    ws = wb.active
    ws.title = "cables"
    headers = ["cable_id", "manufacturer", "type", "conductor", "size",
               "max_amps", "max_temp_f", "source"]

    def cable_id(c: dict) -> str:
        return (f"{c['manufacturer']}_{c['type']}_{c['conductor']}"
                f"_{c['size']}").replace(" ", "")

    rows = [[cable_id(c), c["manufacturer"], c["type"], c["conductor"],
             c["size"], c["max_amps"], c["max_temp_f"],
             c.get("_source", "")] for c in cables]
    write_sheet(ws, headers, rows)

    ws2 = wb.create_sheet("voltage_drop")
    headers2 = ["cable_id", "temp_f", "v_per_amp_per_1000ft"]
    rows2 = []
    for c in cables:
        cid = cable_id(c)
        for t, v in sorted(c["voltage_drop_v_per_amp_per_1000ft"].items(),
                           key=lambda kv: float(kv[0])):
            rows2.append([cid, int(t), v])
    write_sheet(ws2, headers2, rows2)

    write_readme(wb, "Catálogo de cables de potencia para BES", [
        ("Hoja 'cables'", "Una fila por tipo/calibre de cable."),
        ("Hoja 'voltage_drop'", "Caída de tensión en función de la temperatura del conductor. Una fila por (cable, temperatura)."),
        ("cable_id", "Clave única legible: fabricante_tipo_conductor_calibre. El conductor (CU/AL) es parte de la clave porque existe el mismo calibre en cobre y aluminio."),
        ("conductor / size", "Material del conductor y calibre AWG."),
        ("max_amps", "Corriente máxima admisible [A]."),
        ("max_temp_f", "Temperatura máxima del conductor [°F]."),
        ("temp_f", "Temperatura del conductor para el punto de caída de tensión [°F]."),
        ("v_per_amp_per_1000ft", "Caída de tensión [V por amperio por cada 1000 pies]. El loader interpola linealmente entre temperaturas."),
        ("source", "Referencia de origen del dato."),
    ])
    wb.save(OUTPUT_DIR / "cables.xlsx")


def migrate_flat(json_name, key, out_name, columns, title, descriptions,
                 list_columns=()):
    entries = load_json(json_name, key)
    wb = Workbook()
    ws = wb.active
    ws.title = key
    rows = []
    for e in entries:
        row = []
        for col in columns:
            src = col if col in e else "_" + col
            val = e.get(src)
            if col in list_columns and isinstance(val, list):
                val = ", ".join(str(x) for x in val)
            row.append(val)
        rows.append(row)
    write_sheet(ws, columns, rows)
    write_readme(wb, title, descriptions)
    wb.save(OUTPUT_DIR / out_name)


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    migrate_pumps()
    migrate_cables()

    migrate_flat(
        "motors.json", "motors", "motors.xlsx",
        ["manufacturer", "series", "model", "hp_rating", "voltage",
         "amperage", "length_ft", "max_temp_f", "od_inches", "source"],
        "Catálogo de motores electrosumergibles",
        [("hp_rating", "Potencia nominal [HP] a 60 Hz."),
         ("voltage / amperage", "Tensión nominal [V] y corriente nominal [A] de placa."),
         ("length_ft", "Longitud del motor [pies]."),
         ("max_temp_f", "Temperatura ambiente máxima de operación [°F]."),
         ("od_inches", "Diámetro externo [pulgadas]; la serie indica el diámetro nominal."),
         ("source", "Referencia de origen del dato.")],
    )

    migrate_flat(
        "seals.json", "seals", "seals.xlsx",
        ["manufacturer", "series", "model", "type", "compatible_motor_series",
         "od_inches", "length_ft", "thrust_capacity_lbs", "max_temp_f",
         "shaft_hp_standard", "shaft_hp_high_strength", "source"],
        "Catálogo de sellos / protectores",
        [("type", "Tipo de sello: labyrinth (laberinto), bag (bolsa) o combined (combinado). Ver Brown §4.5325."),
         ("compatible_motor_series", "Series de motor compatibles, separadas por comas."),
         ("thrust_capacity_lbs", "Capacidad del cojinete de empuje axial [lbs]."),
         ("shaft_hp_standard / shaft_hp_high_strength", "Potencia máxima transmisible por el eje estándar / de alta resistencia [HP]."),
         ("source", "Referencia de origen del dato.")],
        list_columns=("compatible_motor_series",),
    )

    migrate_flat(
        "gas_handlers.json", "gas_handlers", "gas_handlers.xlsx",
        ["manufacturer", "series", "model", "type", "position", "od_inches",
         "length_ft", "weight_lbs", "hp", "min_flow_bpd", "max_flow_bpd",
         "max_efficiency", "source", "range_source"],
        "Catálogo de separadores / manejadores de gas",
        [("type", "Tipo: vortex, rotary, etc. Los vortex tienen la mayor eficiencia de separación."),
         ("position", "Posición de instalación en el ensamble."),
         ("hp", "Potencia consumida [HP]."),
         ("min_flow_bpd / max_flow_bpd", "Rango de caudal operativo [bpd]."),
         ("max_efficiency", "Eficiencia máxima de separación de gas [fracción 0-1]."),
         ("source / range_source", "Referencias de origen del equipo y de su rango operativo.")],
    )

    migrate_flat(
        "sensors.json", "sensors", "sensors.xlsx",
        ["manufacturer", "model", "intake_pressure_max_psi",
         "intake_temp_max_f", "discharge_pressure_max_psi",
         "motor_winding_temp_max_f", "vibration_monitoring",
         "vibration_max_g", "od_inches", "length_in", "weight_lbs",
         "max_motor_voltage", "source"],
        "Catálogo de sensores de fondo",
        [("intake_pressure_max_psi", "Presión de admisión máxima medible [psi]."),
         ("intake_temp_max_f", "Temperatura de admisión máxima [°F]."),
         ("discharge_pressure_max_psi", "Presión de descarga máxima [psi]; vacío = el sensor no la mide."),
         ("vibration_monitoring / vibration_max_g", "Si mide vibración y su rango [g]; vacío = no aplica."),
         ("max_motor_voltage", "Tensión máxima del motor compatible [V]."),
         ("source", "Referencia de origen del dato.")],
    )

    print(f"OK - archivos generados en {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
