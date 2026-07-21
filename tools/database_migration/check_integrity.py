"""
Chequeo de integridad de la base de datos Excel (esquema v3)
============================================================

Hallazgo H11 de la auditoría: Excel no tiene motor que refuerce claves ni
restricciones — este script ES las "CHECK constraints" de la etapa Excel.
Valida lo que los README documentan:

* unicidad de claves primarias (simples y compuestas),
* integridad referencial (FK sin destino = error),
* restricciones de rango (min<bep<max, 0<eficiencia<1, valores > 0),
* relaciones 1:1 de los pozos (cada well_id presente en cada bloque).

Debe ejecutarse después de cada edición manual de los Excel y antes de
cada integración. En SQLite, estas verificaciones pasan a ser PRIMARY KEY,
FOREIGN KEY y CHECK reforzadas por el motor.

USO:  python check_integrity.py       (exit 0 = OK, exit 1 = errores)
"""
# v3.1: bombas de catálogo sin curva digitalizada generan AVISO, no error.
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).parent / "data_excel"
errors: list[str] = []
warnings: list[str] = []


def read(fname: str, sheet: str) -> list[dict]:
    wb = load_workbook(BASE / fname, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    out = [dict(zip(headers, r)) for r in rows
           if not all(v is None for v in r)]
    wb.close()
    return out


def check_pk(label: str, rows: list[dict], cols: tuple[str, ...]) -> None:
    seen = set()
    for r in rows:
        key = tuple(str(r.get(c)) for c in cols)
        if any(v in ("None", "") for v in key):
            errors.append(f"{label}: PK con celda vacía: {key}")
        elif key in seen:
            errors.append(f"{label}: PK duplicada: {key}")
        seen.add(key)


def check_fk(label: str, rows: list[dict], col, target: set,
             target_name: str, optional: bool = False) -> None:
    cols = col if isinstance(col, tuple) else (col,)
    for r in rows:
        vals = tuple(str(r.get(c)) for c in cols)
        if all(v in ("None", "") for v in vals):
            if not optional:
                errors.append(f"{label}: FK {cols} vacía")
            continue
        key = vals if len(vals) > 1 else vals[0]
        if key not in target:
            errors.append(f"{label}: FK {key!r} sin destino en {target_name}")


def main() -> None:
    # ---- tablas maestras ------------------------------------------------
    manufacturers = read("manufacturers.xlsx", "manufacturers")
    series = read("equipment_series.xlsx", "equipment_series")
    sources = read("data_sources.xlsx", "data_sources")
    m_ids = {str(r["manufacturer_id"]) for r in manufacturers}
    s_ids = {str(r["series_id"]) for r in series}
    src_ids = {str(r["source_id"]) for r in sources}
    check_pk("manufacturers", manufacturers, ("manufacturer_id",))
    check_pk("equipment_series", series, ("series_id",))
    check_pk("data_sources", sources, ("source_id",))

    # ---- bombas ---------------------------------------------------------
    pumps = read("pumps.xlsx", "pumps")
    curves = read("pumps.xlsx", "pump_curves")
    housings = read("pumps.xlsx", "pump_housings")
    p_ids = {str(r["pump_id"]) for r in pumps}
    check_pk("pumps", pumps, ("pump_id",))
    check_pk("pump_curves", curves, ("pump_id", "flow_bpd"))
    check_pk("pump_housings", housings, ("pump_id", "stages"))
    check_fk("pumps", pumps, "manufacturer_id", m_ids, "manufacturers")
    check_fk("pumps", pumps, "series", s_ids, "equipment_series")
    check_fk("pumps", pumps, "source_id", src_ids, "data_sources")
    check_fk("pump_curves", curves, "pump_id", p_ids, "pumps")
    check_fk("pump_housings", housings, "pump_id", p_ids, "pumps")
    for r in pumps:
        pid = r["pump_id"]
        if not (0 < r["min_flow_bpd"] < r["bep_flow_bpd"] < r["max_flow_bpd"]):
            errors.append(f"pumps:{pid}: no cumple 0 < min < bep < max")
        if not r["od_inches"] > 0 or not r["max_stages"] > 0:
            errors.append(f"pumps:{pid}: od o max_stages no positivos")
        if pid not in {c["pump_id"] for c in curves}:
            # v3.1: curva no digitalizada = aviso (la bomba existe en el
            # catálogo pero no está disponible para diseño todavía)
            warnings.append(f"pumps:{pid}: sin curva digitalizada — "
                            f"no disponible para diseño")
    for r in curves:
        if not (0 < r["efficiency"] < 1):
            errors.append(f"pump_curves:{r['pump_id']}@{r['flow_bpd']}: "
                          f"eficiencia fuera de (0,1)")
        if r["head_ft_per_stage"] < 0 or r["hp_per_stage"] <= 0:
            errors.append(f"pump_curves:{r['pump_id']}@{r['flow_bpd']}: "
                          f"head/hp inválidos")

    # ---- motores ----------------------------------------------------------
    motors = read("motors.xlsx", "motors")
    check_pk("motors", motors, ("motor_id",))
    check_fk("motors", motors, "manufacturer_id", m_ids, "manufacturers")
    check_fk("motors", motors, "series", s_ids, "equipment_series")
    check_fk("motors", motors, "source_id", src_ids, "data_sources")
    for r in motors:
        if not (r["hp_rating"] > 0 and r["voltage"] > 0 and r["amperage"] > 0):
            errors.append(f"motors:{r['motor_id']}: hp/V/A no positivos")

    # ---- cables -----------------------------------------------------------
    cables = read("cables.xlsx", "cables")
    vdrop = read("cables.xlsx", "conductor_voltage_drop")
    check_pk("cables", cables, ("cable_id",))
    check_pk("conductor_voltage_drop", vdrop, ("conductor", "size", "temp_f"))
    check_fk("cables", cables, "manufacturer_id", m_ids, "manufacturers")
    check_fk("cables", cables, "source_id", src_ids, "data_sources")
    vd_keys = {(str(r["conductor"]), str(r["size"])) for r in vdrop}
    check_fk("cables", cables, ("conductor", "size"), vd_keys,
             "conductor_voltage_drop")
    for r in cables:
        if not r["max_amps"] > 0:
            errors.append(f"cables:{r['cable_id']}: max_amps no positivo")

    # ---- sellos -----------------------------------------------------------
    seals = read("seals.xlsx", "seals")
    compat = read("seals.xlsx", "seal_motor_compatibility")
    seal_ids = {str(r["seal_id"]) for r in seals}
    check_pk("seals", seals, ("seal_id",))
    check_pk("seal_motor_compatibility", compat, ("seal_id", "motor_series"))
    check_fk("seals", seals, "manufacturer_id", m_ids, "manufacturers")
    check_fk("seals", seals, "series", s_ids, "equipment_series")
    check_fk("seals", seals, "source_id", src_ids, "data_sources")
    check_fk("seal_motor_compatibility", compat, "seal_id", seal_ids, "seals")
    check_fk("seal_motor_compatibility", compat, "motor_series", s_ids,
             "equipment_series")
    motor_series = {str(r["series"]) for r in motors}
    for serie in sorted({str(r["motor_series"]) for r in compat}
                        - motor_series):
        warnings.append(
            f"seal_motor_compatibility: la serie {serie} no tiene motores "
            f"cargados aún (serie válida, catálogo incompleto — H2)")

    # ---- resto de catálogos -------------------------------------------------
    ghs = read("gas_handlers.xlsx", "gas_handlers")
    check_pk("gas_handlers", ghs, ("gas_handler_id",))
    check_fk("gas_handlers", ghs, "manufacturer_id", m_ids, "manufacturers")
    check_fk("gas_handlers", ghs, "source_id", src_ids, "data_sources")
    check_fk("gas_handlers", ghs, "range_source_id", src_ids, "data_sources",
             optional=True)
    for r in ghs:
        if not r["min_flow_bpd"] < r["max_flow_bpd"]:
            errors.append(f"gas_handlers:{r['gas_handler_id']}: min >= max")
        if r["max_efficiency"] is not None and not (0 < r["max_efficiency"] <= 1):
            errors.append(f"gas_handlers:{r['gas_handler_id']}: "
                          f"eficiencia fuera de (0,1]")

    sensors = read("sensors.xlsx", "sensors")
    check_pk("sensors", sensors, ("sensor_id",))
    check_fk("sensors", sensors, "manufacturer_id", m_ids, "manufacturers")
    check_fk("sensors", sensors, "source_id", src_ids, "data_sources")

    transformers = read("transformers.xlsx", "transformers")
    check_pk("transformers", transformers, ("transformer_id",))
    check_fk("transformers", transformers, "manufacturer_id", m_ids,
             "manufacturers")
    check_fk("transformers", transformers, "source_id", src_ids,
             "data_sources")
    for r in transformers:
        if not r["kva_rating"] > 0:
            errors.append(f"transformers:{r['transformer_id']}: kVA <= 0")
        if r["phases"] not in (1, 3):
            errors.append(f"transformers:{r['transformer_id']}: phases ∉ {{1,3}}")

    # ---- pozos de ejemplo: relaciones 1:1 -----------------------------------
    wells = read("well_examples.xlsx", "wells")
    w_ids = {str(r["well_id"]) for r in wells}
    check_pk("wells(examples)", wells, ("well_id",))
    for sheet in ("reservoir", "fluid", "well_geometry",
                  "surface_conditions", "design_objectives",
                  "book_reference"):
        block = read("well_examples.xlsx", sheet)
        check_pk(f"well_examples/{sheet}", block, ("well_id",))
        check_fk(f"well_examples/{sheet}", block, "well_id", w_ids, "wells")
        missing = w_ids - {str(r["well_id"]) for r in block}
        for wid in sorted(missing):
            errors.append(f"well_examples/{sheet}: falta la fila 1:1 "
                          f"del pozo {wid}")
    details = read("well_examples.xlsx", "book_reference_details")
    check_pk("book_reference_details", details,
             ("well_id", "detail_group", "item"))
    check_fk("book_reference_details", details, "well_id", w_ids, "wells")

    # ---- resultado ------------------------------------------------------------
    for w in warnings:
        print("AVISO :", w)
    if errors:
        print(f"\nINTEGRIDAD FALLÓ: {len(errors)} errores")
        for e in errors[:50]:
            print(" -", e)
        sys.exit(1)
    n_tables = 14
    print(f"\nINTEGRIDAD OK: claves primarias únicas, todas las FK con "
          f"destino, restricciones de rango cumplidas y relaciones 1:1 "
          f"completas en los {n_tables} archivos "
          f"({len(warnings)} avisos, 0 errores).")


if __name__ == "__main__":
    main()
