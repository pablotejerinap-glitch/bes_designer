"""
Extractor de curvas de rendimiento — catálogo Alkhorayef 2019 (final)
=====================================================================
Digitaliza Head/Power/Efficiency de los gráficos del catálogo (imágenes
JPEG embebidas): OCR de ticks (tesseract, con fallback por segmentos) +
ajuste robusto RANSAC + extracción por máscara de color + QA por bomba
(pico de eficiencia en el BEP, identidad hidráulica eff=Q·H/(135773·HP)
dentro de ±12%, head decreciente). Las que no pasan el QA se reportan
para lectura visual. Idempotente. La leyenda del catálogo tiene los
colores intercambiados; asignación por color de eje y física (ALKH-02).

USO: python extract_curves_alkhorayef.py <pdf> [--solo M] [--skip N]
     [--take M] [--force-bep]

Resultado de la corrida completa: 35/37 bombas automáticas (WD-3000 con
--force-bep: el pico real de la curva del catálogo no coincide con el
BEP declarado en su ficha) + 2 por lectura visual (WE-8500, WN-1050,
ticks de caudal ilegibles para OCR).
"""
from __future__ import annotations
import re, subprocess, sys, tempfile
from pathlib import Path
import numpy as np
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font

DATA_DIR = Path(__file__).parent / "data_excel"
HYD_K = 135_773.0
MANUFACTURER = "Alkhorayef"
SOURCE_ID = "ALKH-02"
CITATION = ("Curvas digitalizadas de los gráficos del Alkhorayef Product Catalog 2019 "
            "(imagen embebida; calibración por OCR de ticks con ajuste robusto + extracción por color; "
            "exactitud estimada ±2%). Leyenda del catálogo con colores intercambiados; asignación por "
            "color de eje y física. Puntos a 60 Hz, SG=1.0.")
CELL_FONT = Font(name="Arial")
ZOOM = 4

def sh(cmd):
    return subprocess.run(cmd, capture_output=True, text=True, check=True).stdout

def find_pump_pages(pdf, n_pages=46):
    mapping = {}
    for p in range(1, n_pages + 1):
        try: txt = sh(["pdftotext", "-f", str(p), "-l", str(p), str(pdf), "-"])
        except subprocess.CalledProcessError: continue
        if "Optimum operating range" not in txt: continue
        m = re.search(r"^(W[A-Z]-\d+)\s*$", txt, re.M)
        if m: mapping[m.group(1)] = p
    return mapping

def chart_image(pdf, page, tmpdir):
    sh(["pdfimages", "-f", str(page), "-l", str(page), "-png", str(pdf), f"{tmpdir}/img"])
    for f in sorted(Path(tmpdir).glob("img-*.png"), key=lambda f: f.stat().st_size, reverse=True):
        im = Image.open(f)
        im.load()
        im = im.convert("RGB")
        if im.width > 400 and im.height > 300: return im
    raise ValueError("sin imagen de gráfico")

def axis_lines(img):
    """Líneas de eje: x0 (izq), x_right (borde del plot), y_top, y_bot."""
    a = np.asarray(img.convert("L"))
    dark = a < 130
    h, w = a.shape
    colscore = dark.sum(axis=0)
    cands = np.nonzero(colscore > h * 0.5)[0]
    if not len(cands): raise ValueError("sin ejes verticales")
    # agrupa columnas contiguas
    groups, cur = [], [cands[0]]
    for c in cands[1:]:
        if c - cur[-1] <= 2: cur.append(c)
        else: groups.append(cur); cur = [c]
    groups.append(cur)
    xs = [int(np.mean(g)) for g in groups]
    x0 = xs[0]
    # el borde derecho es la ÚLTIMA línea vertical larga: las líneas
    # intermedias son los divisores Compression/Floater del catálogo
    x_right = xs[-1] if len(xs) > 1 else w - 20
    col = dark[:, x0]
    ys = np.nonzero(col)[0]
    return x0, x_right, int(ys.min()), int(ys.max())

def ocr_numbers(img):
    big = img.resize((img.width*ZOOM, img.height*ZOOM), Image.LANCZOS)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        big.save(f.name)
        tsv = subprocess.run(["tesseract", f.name, "-", "--psm", "11", "-c",
            "tessedit_char_whitelist=0123456789.,", "tsv"], capture_output=True, text=True).stdout
    Path(f.name).unlink()
    out = []
    for line in tsv.splitlines()[1:]:
        p = line.split("\t")
        if len(p) < 12 or not p[11].strip(): continue
        t = p[11].strip().strip(".,")
        if not re.fullmatch(r"[\d.,]+", t) or len(t) > 7: continue
        t = t.replace(",", "")
        try: v = float(t)
        except ValueError: continue
        x, y, w_, h_ = (int(u)/ZOOM for u in p[6:10])
        out.append({"x": x + w_/2, "y": y + h_/2, "v": v, "t": t})
    return out


def xticks_by_segments(img, geo):
    """OCR de la banda inferior segmento a segmento (fallback)."""
    import numpy as _np, tempfile as _tf, subprocess as _sp, re as _re
    from pathlib import Path as _P
    x0, x_right, y_top, y_bot = geo
    rgb = _np.asarray(img.convert("RGB")).astype(int)
    mx = rgb.max(axis=2); mn = rgb.min(axis=2)
    text = (mx < 165) & (mx - mn < 45)          # texto negro/gris, no curvas de color
    y0s = max(y_bot - 16, 0); y1s = min(y_bot + 20, text.shape[0])
    strip = text[y0s:y1s, :].copy()
    w = strip.shape[1]
    for i in range(strip.shape[0]):             # borra la línea del eje y ticks largos
        if strip[i].sum() > w * 0.35: strip[i] = False
    dark = strip.sum(axis=0) > 0
    # segmentos: columnas oscuras, uniendo huecos internos (< 8 px)
    segs, cur, gap = [], None, 0
    for x, d in enumerate(dark):
        if d:
            if cur is None: cur = x
            gap = 0
        elif cur is not None:
            gap += 1
            if gap >= 8:
                if x - gap - cur >= 4: segs.append((cur, x - gap + 1))
                cur, gap = None, 0
    if cur is not None: segs.append((cur, len(dark)))
    out = []
    for sx, ex in segs:
        crop = img.crop((max(sx-3,0), y0s, min(ex+3, img.width), y1s))
        big = crop.resize((crop.width*6, crop.height*6))
        with _tf.NamedTemporaryFile(suffix=".png", delete=False) as f:
            big.save(f.name)
            txt = _sp.run(["tesseract", f.name, "-", "--psm", "8", "-c",
                "tessedit_char_whitelist=0123456789,"], capture_output=True, text=True).stdout.strip()
        _P(f.name).unlink()
        txt = txt.replace(",", "").strip()
        if _re.fullmatch(r"\d{1,6}", txt):
            out.append(((sx+ex)/2, float(txt)))
    return out

def fit_axis_ransac(pairs, min_inliers=2, sign=0):
    """pairs=[(px, valor)] → función lineal robusta o None."""
    pairs = [p for p in pairs if p[1] < 1e6]
    n = len(pairs)
    if n < 2: return None
    vmax = max(p[1] for p in pairs); vmin = min(p[1] for p in pairs)
    tol = max(0.04 * (vmax - vmin), 0.35)
    best = None
    for i in range(n):
        for j in range(i+1, n):
            (p1, v1), (p2, v2) = pairs[i], pairs[j]
            if abs(p1 - p2) < 4 or v1 == v2: continue
            m = (v2 - v1) / (p2 - p1)
            if sign and m * sign <= 0: continue
            b = v1 - m * p1
            inl = [(p, v) for p, v in pairs if abs(m*p + b - v) <= tol]
            if best is None or len(inl) > len(best):
                best = inl
    if best is None or len(best) < min_inliers: return None
    px = np.array([p[0] for p in best]); vv = np.array([p[1] for p in best])
    if len(best) == 2:
        m = (vv[1]-vv[0])/(px[1]-px[0]); A = (m, vv[0]-m*px[0])
    else:
        A = np.polyfit(px, vv, 1)
    return (lambda p: A[0]*p + A[1]), len(best)

CURRENT_IMG = None

def calibrate(nums, geo, img_w):
    x0, x_right, y_top, y_bot = geo
    xt = [(n["x"], n["v"]) for n in nums if n["y"] > y_bot - 18 and "." not in n["t"] and x0 - 20 <= n["x"] <= x_right + 20]
    x_axis = fit_axis_ransac(xt, sign=1)
    if x_axis is None and globals().get('CURRENT_IMG') is not None:
        x_axis = fit_axis_ransac(xticks_by_segments(globals()['CURRENT_IMG'], (x0, 0, 0, y_bot)), sign=1)
    ht = [(n["y"], n["v"]) for n in nums if n["x"] < x0 - 2 and y_top-8 <= n["y"] <= y_bot-12]
    head_axis = fit_axis_ransac(ht, min_inliers=3, sign=-1)
    right = [n for n in nums if n["x"] > x0 + 0.65 * (img_w - x0) and y_top+5 <= n["y"] <= y_bot-12]
    if right:
        rx = sorted(set(round(n["x"]) for n in right))
        gaps = [(rx[i+1]-rx[i], i) for i in range(len(rx)-1)]
        split = None
        if gaps:
            gmax, gi = max(gaps)
            if gmax > 15: split = (rx[gi] + rx[gi+1]) / 2
        inner = [n for n in right if split is None or n["x"] < split]
        outer = [n for n in right if split is not None and n["x"] >= split]
    else:
        inner, outer = [], []
    # power = clúster interior (enteros o decimales; '044'→0.44 si el
    # resto del clúster es decimal); eficiencia = clúster exterior (0-100)
    def fix_zero(n):
        if "." not in n["t"] and n["t"].startswith("0") and len(n["t"]) > 1:
            return float("0." + n["t"][1:])
        return n["v"]
    if outer:
        pt = [(n["y"], fix_zero(n)) for n in inner]
        et = [(n["y"], n["v"]) for n in outer if "." not in n["t"] and n["v"] <= 100]
    else:
        pt = [(n["y"], fix_zero(n)) for n in inner if "." in n["t"] or n["v"] < 1
              or (n["t"].startswith("0") and len(n["t"]) > 1)]
        et = [(n["y"], n["v"]) for n in inner if "." not in n["t"] and 1 <= n["v"] <= 100]
    power_axis = fit_axis_ransac(pt, min_inliers=3, sign=-1)
    eff_axis = fit_axis_ransac(et, min_inliers=3, sign=-1)
    missing = [nm for nm, a in [("caudal", x_axis), ("head", head_axis),
               ("power", power_axis), ("eficiencia", eff_axis)] if a is None]
    if missing: raise ValueError(f"faltan ejes {missing}")
    return x_axis[0], head_axis[0], power_axis[0], eff_axis[0]

def extract_curves(img, geo):
    x0, x_right, y_top, y_bot = geo
    a = np.asarray(img).astype(int)
    R, G, B = a[...,0], a[...,1], a[...,2]
    masks = {"blue": (B>120)&(B-R>45)&(B-G>45),
             "green": (G>90)&(G-R>35)&(G-B>35),
             "red": (R>130)&(R-G>45)&(R-B>45)}
    curves = {}
    for name, mask in masks.items():
        m = mask.copy()
        m[:y_top+2,:] = False; m[y_bot-1:,:] = False
        m[:,:x0+2] = False; m[:,x_right-1:] = False
        ys_img, xs_img = np.nonzero(m)
        d = {}
        for x in np.unique(xs_img):
            d[int(x)] = float(np.median(ys_img[xs_img == x]))
        curves[name] = d
    return curves

def sample(curve, x_axis, y_axis, q_points):
    out = []
    xs_sorted = np.array(sorted(curve))
    if not len(xs_sorted): return [None]*len(q_points)
    qs = np.array([x_axis(x) for x in xs_sorted])
    span = (qs.max()-qs.min())/max(len(qs)-1,1)
    for q in q_points:
        i = int(np.argmin(np.abs(qs-q)))
        if abs(qs[i]-q) > max(span*6, 0.02*abs(q)+5):
            out.append(None); continue
        sel = np.abs(qs-qs[i]) <= span*3
        out.append(y_axis(float(np.mean([curve[int(x)] for x in xs_sorted[sel]]))))
    return out

def qa(q_points, heads, powers, effs, bep):
    problems = []
    pts = [(q,h,p,e) for q,h,p,e in zip(q_points,heads,powers,effs) if None not in (h,p,e)]
    if len(pts) < 7:
        return [f"solo {len(pts)} puntos válidos"]
    q_pk = max(pts, key=lambda p: p[3])[0]
    if abs(q_pk-bep) > 0.12*bep: problems.append(f"pico eff en {q_pk:.0f} (BEP {bep:.0f})")
    bad = sum(1 for q,h,p,e in pts if e >= 0.05 and p > 0
              and abs(q*h/(HYD_K*p) - e)/e > 0.12)
    if bad > len(pts)*0.34: problems.append(f"identidad hidráulica falla {bad}/{len(pts)}")
    qs = [p[0] for p in pts]; qh = 0.25*max(qs)
    ha = [(q,h) for q,h,_,_ in pts if q > qh]
    for (_,h1),(qb,h2) in zip(ha, ha[1:]):
        if h2 > h1*1.03: problems.append(f"head sube en {qb:.0f}"); break
    return problems

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    solo = sys.argv[sys.argv.index("--solo")+1] if "--solo" in sys.argv else None
    skip = int(sys.argv[sys.argv.index("--skip")+1]) if "--skip" in sys.argv else 0
    take = int(sys.argv[sys.argv.index("--take")+1]) if "--take" in sys.argv else 999
    pdf = Path(args[0])
    wbp = load_workbook(DATA_DIR/"pumps.xlsx")
    wsp = wbp["pumps"]
    H = {c.value: i+1 for i,c in enumerate(wsp[1]) if c.value}
    fichas = {}
    for r in range(2, wsp.max_row+1):
        if wsp.cell(row=r, column=H["manufacturer_id"]).value == MANUFACTURER:
            mo = str(wsp.cell(row=r, column=H["model"]).value)
            fichas[mo] = {"row": r, "pump_id": wsp.cell(row=r, column=H["pump_id"]).value,
                "bep": float(wsp.cell(row=r, column=H["bep_flow_bpd"]).value),
                "qmin": float(wsp.cell(row=r, column=H["min_flow_bpd"]).value),
                "qmax": float(wsp.cell(row=r, column=H["max_flow_bpd"]).value)}
    pages = find_pump_pages(pdf)
    ok_models, review = {}, {}
    for model, page in sorted(pages.items())[skip:skip+take]:
        if solo and model != solo: continue
        if model not in fichas: review[model] = ["sin ficha"]; continue
        f = fichas[model]
        try:
            with tempfile.TemporaryDirectory() as td:
                img = chart_image(pdf, page, td)
            geo = axis_lines(img)
            globals()['CURRENT_IMG'] = img
            ns_dummy = None
            x_axis, head_ax, pow_ax, eff_ax = calibrate(ocr_numbers(img), geo, img.width)
            curves = extract_curves(img, geo)
            blue, green, red = curves["blue"], curves["green"], curves["red"]
            if not blue or not green or not red: raise ValueError("curva sin píxeles")
            q_max_curve = max(x_axis(x) for x in blue)
            q_lo = max(f["qmin"]*0.6, 0.05*q_max_curve)
            q_hi = min(f["qmax"]*1.15, q_max_curve*0.98)
            q_points = sorted(set([float(round(q)) for q in np.linspace(q_lo,q_hi,9)] + [f["bep"]]))
            heads = sample(blue, x_axis, head_ax, q_points)
            powers = sample(red, x_axis, pow_ax, q_points)
            effs = [e/100.0 if e is not None else None for e in sample(green, x_axis, eff_ax, q_points)]
            problems = qa(q_points, heads, powers, effs, f["bep"])
            if "--force-bep" in sys.argv:
                problems = [p for p in problems if "pico eff" not in str(p)]
            near0 = [y for x,y in blue.items() if 0 <= x_axis(x) <= 0.02*q_max_curve]
            shutoff = head_ax(float(np.median(near0))) if near0 else None
            if problems: review[model] = problems
            else:
                pts = [(q,h,p,e) for q,h,p,e in zip(q_points,heads,powers,effs)
                       if None not in (h,p,e) and h > 0 and e > 0.005]
                ok_models[model] = {"points": pts, "shutoff": shutoff,
                                    "pump_id": f["pump_id"], "row": f["row"]}
        except Exception as exc:
            review[model] = [f"error: {exc}"]
    ws_c = wbp["pump_curves"]
    ch = {c.value: i+1 for i,c in enumerate(ws_c[1]) if c.value}
    done_ids = {d["pump_id"] for d in ok_models.values()}
    for r in reversed([r for r in range(2, ws_c.max_row+1)
                       if ws_c.cell(row=r, column=ch["pump_id"]).value in done_ids]):
        ws_c.delete_rows(r)
    if "shutoff_head_ft_per_stage" not in H:
        col = len(H)+1
        c = wsp.cell(row=1, column=col, value="shutoff_head_ft_per_stage")
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        H["shutoff_head_ft_per_stage"] = col
    n_pts = 0
    for model, data in ok_models.items():
        for q,h,p,e in data["points"]:
            r = ws_c.max_row+1
            vals = {"pump_id": data["pump_id"], "flow_bpd": round(q),
                    "head_ft_per_stage": round(h,2), "hp_per_stage": round(p,4),
                    "efficiency": round(e,3)}
            for cn, idx in ch.items():
                ws_c.cell(row=r, column=idx, value=vals.get(cn)).font = CELL_FONT
            n_pts += 1
        if data["shutoff"]:
            wsp.cell(row=data["row"], column=H["shutoff_head_ft_per_stage"],
                     value=round(data["shutoff"],2)).font = CELL_FONT
    wbp.save(DATA_DIR/"pumps.xlsx")
    wb = load_workbook(DATA_DIR/"data_sources.xlsx")
    ws = wb["data_sources"]
    sm = {c.value: i+1 for i,c in enumerate(ws[1]) if c.value}
    if not any(str(ws.cell(row=r, column=sm["source_id"]).value) == SOURCE_ID
               for r in range(2, ws.max_row+1)):
        r = ws.max_row+1
        for cn, v in [("source_id", SOURCE_ID), ("type", "digitalización de gráfico"),
                      ("citation", CITATION)]:
            ws.cell(row=r, column=sm[cn], value=v).font = CELL_FONT
        wb.save(DATA_DIR/"data_sources.xlsx")
    print(f"CARGADAS: {len(ok_models)} bombas, {n_pts} puntos")
    if review:
        print(f"PARA REVISIÓN ({len(review)}):")
        for m, probs in sorted(review.items()):
            print(f"  {m}: {'; '.join(str(x) for x in probs)}")


if __name__ == "__main__":
    main()
