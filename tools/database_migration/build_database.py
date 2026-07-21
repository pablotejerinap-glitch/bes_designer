"""
Constructor de la base de datos Excel — BES Designer (esquema v2)
=================================================================

Genera en ``data_excel/`` la base de datos completa y normalizada descrita
en ``DISENO_BASE_DE_DATOS.md``:

* catálogos migrados desde ``catalogs/*.json`` (bombas, motores, cables,
  sellos, separadores, sensores),
* tabla ``manufacturers`` (3FN: el fabricante deja de repetirse como texto),
* listas convertidas a tablas de detalle (1FN): pump_housings,
  seal_motor_compatibility,
* ``transformers`` extraída de los datos hardcodeados en core/electrical.py,
* ``vsds`` (plantilla vacía para catálogos futuros),
* ``well_examples`` migrada desde data/example_wells.json,
* ``real_wells`` (plantilla para casos de campo).

Cada workbook incluye una hoja README con claves primarias, foráneas,
relaciones, unidades y restricciones — en español.

Supersede a ``migrate_json_to_excel.py`` (esquema v1, sin normalizar).
No modifica el proyecto original: solo lee JSON y escribe en data_excel/.

USO:  python build_database.py
"""
from __future__ import annotations

import json
from pathlib import Path

import bes

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
# tools/database_migration/ -> raiz del repo
_REPO_ROOT = BASE_DIR.parent.parent
# Los catalogos JSON son datos del paquete instalable: se resuelven a
# traves de el, no por ruta relativa al repo.
_BES_PKG = Path(bes.__file__).parent
CATALOG_DIR = _BES_PKG / "catalogs"
DATA_DIR = _REPO_ROOT / "backend" / "data"
OUTPUT_DIR = BASE_DIR / "data_excel"

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
CELL_FONT = Font(name=FONT)

# Tamaños estándar de transformador trifásico [kVA] — copiados de
# core/electrical.py::_TRANSFORMER_SIZES_KVA. Al integrar la base de
# datos, electrical.py leerá ESTA tabla y la tupla del código se retira.
TRANSFORMER_SIZES_KVA = (25.0, 37.5, 50.0, 75.0, 100.0, 150.0, 200.0, 300.0)
TRANSFORMER_LOSS_PCT = 2.5  # pérdida típica usada por calculate_surface_voltage()

# Información conocida de fabricantes (completar a medida que se agreguen)
MANUFACTURER_INFO = {
    "Reda": ("Schlumberger REDA (SLB)", "EE.UU."),
    "Centrilift": ("Baker Hughes Centrilift", "EE.UU."),
    "SLB": ("SLB (Schlumberger)", "EE.UU."),
    "Weatherford": ("Weatherford International", "EE.UU."),
    "ChampionX": ("ChampionX Artificial Lift", "EE.UU."),
    "Generic": ("Tamaños estándar de norma (sin fabricante)", ""),
}


# ----------------------------------------------------------------------
# Helpers de escritura
# ----------------------------------------------------------------------
def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = CELL_FONT
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(width + 3, 60)
    ws.freeze_panes = "A2"


def write_readme(wb, title, sections):
    """Hoja README. ``sections`` = lista de (encabezado, [(item, desc), ...])."""
    ws = wb.create_sheet("README")
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    r = 3
    for heading, items in sections:
        ws[f"A{r}"] = heading
        ws[f"A{r}"].font = Font(name=FONT, bold=True, size=12, color="1F4E79")
        r += 1
        for item, desc in items:
            ws[f"A{r}"] = item
            ws[f"B{r}"] = desc
            ws[f"A{r}"].font = Font(name=FONT, bold=True)
            ws[f"B{r}"].font = CELL_FONT
            ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 105


def load_json(directory, name, key):
    with open(directory / name, encoding="utf-8") as f:
        return json.load(f)[key]


def make_id(entry):
    parts = [entry["manufacturer"], entry.get("series", ""), entry["model"]]
    return "_".join(str(p) for p in parts if p)


def cable_id(c):
    return (f"{c['manufacturer']}_{c['type']}_{c['conductor']}"
            f"_{c['size']}").replace(" ", "")


KEY_NOTE = ("Convención de claves",
            [("PK", "Clave primaria: identifica unívocamente cada fila. En SQLite será PRIMARY KEY."),
             ("FK", "Clave foránea: referencia la PK de otra tabla. En SQLite será FOREIGN KEY."),
             ("Celda vacía", "Equivale a NULL: dato no aplicable o no disponible. Nunca usar 0 como 'sin dato'.")])


# ======================================================================
# manufacturers
# ======================================================================
def build_manufacturers(all_manufacturers):
    wb = Workbook()
    ws = wb.active
    ws.title = "manufacturers"
    rows = []
    for m in sorted(all_manufacturers):
        full, country = MANUFACTURER_INFO.get(m, (m, ""))
        rows.append([m, full, country, ""])
    write_sheet(ws, ["manufacturer_id", "full_name", "country", "notes"], rows)
    write_readme(wb, "Tabla maestra de fabricantes", [
        ("Claves y relaciones",
         [("manufacturer_id", "PK. Nombre corto único del fabricante (ej. 'Reda')."),
          ("Relaciones", "1:N con pumps, motors, cables, seals, gas_handlers, sensors, transformers y vsds: cada equipo referencia manufacturer_id como FK.")]),
        ("Columnas",
         [("full_name", "Razón social o nombre comercial completo."),
          ("country", "País de origen (informativo)."),
          ("notes", "Observaciones (representante local, estado comercial, etc.).")]),
        ("Cómo agregar un fabricante",
         [("Paso único", "Agregar una fila aquí y luego sus equipos en los otros archivos usando el mismo manufacturer_id. No se modifica ningún código.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "manufacturers.xlsx")


# ======================================================================
# pumps: pumps + pump_curves + pump_housings
# ======================================================================
def build_pumps():
    pumps = load_json(CATALOG_DIR, "pumps.json", "pumps")
    wb = Workbook()

    ws = wb.active
    ws.title = "pumps"
    write_sheet(ws,
                ["pump_id", "manufacturer_id", "series", "model", "od_inches",
                 "min_flow_bpd", "max_flow_bpd", "bep_flow_bpd", "max_stages",
                 "source"],
                [[make_id(p), p["manufacturer"], p["series"], p["model"],
                  p["od_inches"], p["min_flow_bpd"], p["max_flow_bpd"],
                  p["bep_flow_bpd"], p["max_stages"], p.get("_source", "")]
                 for p in pumps])

    ws2 = wb.create_sheet("pump_curves")
    write_sheet(ws2,
                ["pump_id", "flow_bpd", "head_ft_per_stage", "hp_per_stage",
                 "efficiency"],
                [[make_id(p), pt["flow_bpd"], pt["head_ft_per_stage"],
                  pt["hp_per_stage"], pt["efficiency"]]
                 for p in pumps for pt in p["performance_curve"]])

    ws3 = wb.create_sheet("pump_housings")
    write_sheet(ws3, ["pump_id", "stages"],
                [[make_id(p), h] for p in pumps for h in p["housing_options"]])

    write_readme(wb, "Catálogo de bombas electrosumergibles (BES/ESP)", [
        ("Tablas, claves y relaciones",
         [("pumps", "Tabla maestra. PK: pump_id (fabricante_serie_modelo). FK: manufacturer_id → manufacturers."),
          ("pump_curves", "Curva de rendimiento del fabricante, punto por punto. PK compuesta: (pump_id, flow_bpd). FK: pump_id → pumps. Relación pumps 1:N pump_curves."),
          ("pump_housings", "Housings disponibles (nº de etapas por housing). PK compuesta: (pump_id, stages). FK: pump_id → pumps. Relación 1:N."),
          ("Integridad", "Si se cambia un pump_id en 'pumps', debe actualizarse en pump_curves y pump_housings (en SQLite esto lo refuerza el motor con FOREIGN KEY).")]),
        ("Columnas de pumps",
         [("od_inches", "Diámetro externo [pulg]. Restricción: > 0 y menor al ID del casing donde se instale."),
          ("min_flow_bpd / max_flow_bpd", "Rango operativo recomendado [bpd]. Restricción: min < max."),
          ("bep_flow_bpd", "Caudal de máxima eficiencia [bpd]. Restricción: min_flow < bep < max_flow."),
          ("max_stages", "Máximo de etapas por limitación de housing/eje. Entero > 0."),
          ("source", "Trazabilidad del dato (Brown Vol. 2B, catálogo del fabricante).")]),
        ("Columnas de pump_curves",
         [("flow_bpd", "Caudal del punto [bpd]. Único por bomba (parte de la PK)."),
          ("head_ft_per_stage", "Altura dinámica por etapa [pies/etapa], con agua (SG=1.0)."),
          ("hp_per_stage", "Potencia al freno por etapa [HP/etapa], con agua; multiplicar por SG del fluido real."),
          ("efficiency", "Eficiencia hidráulica [0–1]. Restricción: 0 < eff < 1."),
          ("Decisión de diseño", "Se guardan los puntos discretos del catálogo (no un polinomio ajustado): cada punto es verificable contra el catálogo impreso y el código interpola linealmente.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "pumps.xlsx")
    return {p["manufacturer"] for p in pumps}


# ======================================================================
# motors
# ======================================================================
def build_motors():
    motors = load_json(CATALOG_DIR, "motors.json", "motors")
    wb = Workbook()
    ws = wb.active
    ws.title = "motors"
    write_sheet(ws,
                ["motor_id", "manufacturer_id", "series", "model", "hp_rating",
                 "voltage", "amperage", "length_ft", "max_temp_f", "od_inches",
                 "source"],
                [[make_id(m), m["manufacturer"], m["series"], m["model"],
                  m["hp_rating"], m["voltage"], m["amperage"], m["length_ft"],
                  m["max_temp_f"], m["od_inches"], m.get("_source", "")]
                 for m in motors])
    write_readme(wb, "Catálogo de motores electrosumergibles", [
        ("Claves y relaciones",
         [("motor_id", "PK: fabricante_serie_modelo (unicidad verificada sobre el catálogo actual)."),
          ("manufacturer_id", "FK → manufacturers. Relación manufacturers 1:N motors."),
          ("series", "Serie = diámetro nominal del frame. La usa seal_motor_compatibility (ver seals.xlsx).")]),
        ("Columnas",
         [("hp_rating", "Potencia nominal [HP] a 60 Hz. Restricción: > 0."),
          ("voltage / amperage", "Tensión [V] y corriente [A] de placa. Un mismo frame aparece una vez por cada combinación V/A que ofrece el fabricante."),
          ("length_ft", "Longitud [pies] (para el ensamble y la maniobra de bajada)."),
          ("max_temp_f", "Temperatura ambiente máxima de operación [°F]. Comparar contra temperatura de fondo."),
          ("od_inches", "Diámetro externo [pulg]."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "motors.xlsx")
    return {m["manufacturer"] for m in motors}


# ======================================================================
# cables + cable_voltage_drop
# ======================================================================
def build_cables():
    cables = load_json(CATALOG_DIR, "cables.json", "cables")
    wb = Workbook()
    ws = wb.active
    ws.title = "cables"
    write_sheet(ws,
                ["cable_id", "manufacturer_id", "type", "conductor", "size",
                 "max_amps", "max_temp_f", "source"],
                [[cable_id(c), c["manufacturer"], c["type"], c["conductor"],
                  c["size"], c["max_amps"], c["max_temp_f"],
                  c.get("_source", "")] for c in cables])

    ws2 = wb.create_sheet("cable_voltage_drop")
    rows2 = []
    for c in cables:
        cid = cable_id(c)
        for t, v in sorted(c["voltage_drop_v_per_amp_per_1000ft"].items(),
                           key=lambda kv: float(kv[0])):
            rows2.append([cid, int(t), v])
    write_sheet(ws2, ["cable_id", "temp_f", "v_per_amp_per_1000ft"], rows2)

    write_readme(wb, "Catálogo de cables de potencia para BES", [
        ("Tablas, claves y relaciones",
         [("cables", "Tabla maestra. PK: cable_id = fabricante_tipo_CONDUCTOR_calibre. FK: manufacturer_id → manufacturers."),
          ("cable_voltage_drop", "PK compuesta: (cable_id, temp_f). FK: cable_id → cables. Relación cables 1:N cable_voltage_drop."),
          ("Por qué el conductor integra la PK", "Existe el mismo tipo y calibre en cobre (CU) y aluminio (AL) con caídas de tensión ~65% distintas. Sin el conductor la clave colisiona — error real detectado por la verificación automática en la etapa 1.")]),
        ("Columnas",
         [("conductor", "Material: CU (cobre) o AL (aluminio)."),
          ("size", "Calibre AWG (#1, #2, #4, #6...)."),
          ("max_amps", "Ampacidad [A]. Restricción: > 0."),
          ("max_temp_f", "Temperatura máxima del conductor [°F]."),
          ("temp_f", "Temperatura del punto de caída de tensión [°F]."),
          ("v_per_amp_per_1000ft", "Caída [V/A/1000 pies]. El código interpola linealmente entre temperaturas."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "cables.xlsx")
    return {c["manufacturer"] for c in cables}


# ======================================================================
# seals + seal_motor_compatibility (N:M)
# ======================================================================
def build_seals():
    seals = load_json(CATALOG_DIR, "seals.json", "seals")
    wb = Workbook()
    ws = wb.active
    ws.title = "seals"
    write_sheet(ws,
                ["seal_id", "manufacturer_id", "series", "model", "type",
                 "od_inches", "length_ft", "thrust_capacity_lbs", "max_temp_f",
                 "shaft_hp_standard", "shaft_hp_high_strength", "source"],
                [[make_id(s), s["manufacturer"], s["series"], s["model"],
                  s["type"], s["od_inches"], s["length_ft"],
                  s["thrust_capacity_lbs"], s["max_temp_f"],
                  s.get("shaft_hp_standard"), s.get("shaft_hp_high_strength"),
                  s.get("_source", "")] for s in seals])

    ws2 = wb.create_sheet("seal_motor_compatibility")
    write_sheet(ws2, ["seal_id", "motor_series"],
                [[make_id(s), serie] for s in seals
                 for serie in s.get("compatible_motor_series", [])])

    write_readme(wb, "Catálogo de sellos / protectores", [
        ("Tablas, claves y relaciones",
         [("seals", "Tabla maestra. PK: seal_id. FK: manufacturer_id → manufacturers."),
          ("seal_motor_compatibility", "Tabla puente de la relación N:M entre sellos y SERIES de motor: un sello sirve para varias series y una serie admite varios sellos. PK compuesta: (seal_id, motor_series). FK: seal_id → seals; motor_series referencia el dominio de motors.series."),
          ("Por qué una tabla puente", "Una relación muchos-a-muchos no puede representarse con una FK simple; la lista separada por comas (esquema v1) violaba la 1ª forma normal e impedía filtrar por serie con una consulta simple.")]),
        ("Columnas de seals",
         [("type", "Tipo de sello: labyrinth (laberinto), bag (bolsa) o combined. Referencia: Brown §4.5325."),
          ("thrust_capacity_lbs", "Capacidad del cojinete de empuje axial [lbs]. Restricción: > 0."),
          ("shaft_hp_standard / shaft_hp_high_strength", "Potencia máxima transmisible por el eje estándar / alta resistencia [HP]. Vacío = no informado."),
          ("max_temp_f", "Temperatura máxima [°F]."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "seals.xlsx")
    return {s["manufacturer"] for s in seals}


# ======================================================================
# gas_handlers y sensors (planos)
# ======================================================================
def build_gas_handlers():
    ghs = load_json(CATALOG_DIR, "gas_handlers.json", "gas_handlers")
    wb = Workbook()
    ws = wb.active
    ws.title = "gas_handlers"
    write_sheet(ws,
                ["gas_handler_id", "manufacturer_id", "series", "model",
                 "type", "position", "od_inches", "length_ft", "weight_lbs",
                 "hp", "min_flow_bpd", "max_flow_bpd", "max_efficiency",
                 "source", "range_source"],
                [[make_id(g), g["manufacturer"], g["series"], g["model"],
                  g["type"], g["position"], g["od_inches"], g["length_ft"],
                  g["weight_lbs"], g["hp"], g["min_flow_bpd"],
                  g["max_flow_bpd"], g["max_efficiency"],
                  g.get("_source", ""), g.get("_range_source", "")]
                 for g in ghs])
    write_readme(wb, "Catálogo de separadores / manejadores de gas", [
        ("Claves y relaciones",
         [("gas_handler_id", "PK: fabricante_serie_modelo."),
          ("manufacturer_id", "FK → manufacturers. Relación 1:N.")]),
        ("Columnas",
         [("type", "vortex, rotary, etc. Los vortex logran la mayor eficiencia de separación."),
          ("position", "Posición en el ensamble (debajo de la bomba, reemplaza la admisión)."),
          ("min_flow_bpd / max_flow_bpd", "Rango operativo [bpd]. Restricción: min < max."),
          ("max_efficiency", "Eficiencia máxima de separación [0–1]."),
          ("source / range_source", "Trazabilidad del equipo y de su rango operativo (pueden tener orígenes distintos).")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "gas_handlers.xlsx")
    return {g["manufacturer"] for g in ghs}


def build_sensors():
    sensors = load_json(CATALOG_DIR, "sensors.json", "sensors")
    wb = Workbook()
    ws = wb.active
    ws.title = "sensors"
    write_sheet(ws,
                ["sensor_id", "manufacturer_id", "model",
                 "intake_pressure_max_psi", "intake_temp_max_f",
                 "discharge_pressure_max_psi", "motor_winding_temp_max_f",
                 "vibration_monitoring", "vibration_max_g", "od_inches",
                 "length_in", "weight_lbs", "max_motor_voltage", "source"],
                [[f"{s['manufacturer']}_{s['model']}".replace(" ", ""),
                  s["manufacturer"], s["model"], s["intake_pressure_max_psi"],
                  s["intake_temp_max_f"], s.get("discharge_pressure_max_psi"),
                  s["motor_winding_temp_max_f"], s["vibration_monitoring"],
                  s.get("vibration_max_g"), s["od_inches"], s["length_in"],
                  s["weight_lbs"], s["max_motor_voltage"],
                  s.get("_source", "")] for s in sensors])
    write_readme(wb, "Catálogo de sensores de fondo", [
        ("Claves y relaciones",
         [("sensor_id", "PK: fabricante_modelo."),
          ("manufacturer_id", "FK → manufacturers. Relación 1:N.")]),
        ("Columnas",
         [("intake_pressure_max_psi / intake_temp_max_f", "Rangos máximos medibles en la admisión [psi] / [°F]."),
          ("discharge_pressure_max_psi", "Vacío (NULL) = el sensor no mide presión de descarga."),
          ("vibration_monitoring / vibration_max_g", "TRUE/FALSE y rango [g]; vacío = no aplica."),
          ("max_motor_voltage", "Tensión máxima del motor compatible [V]."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "sensors.xlsx")
    return {s["manufacturer"] for s in sensors}


# ======================================================================
# transformers (datos extraídos de core/electrical.py) y vsds (plantilla)
# ======================================================================
def build_transformers():
    wb = Workbook()
    ws = wb.active
    ws.title = "transformers"
    src = ("Tamaños estándar trifásicos ANSI/IEEE; extraído de "
           "core/electrical.py::_TRANSFORMER_SIZES_KVA (esquema v2)")
    rows = [[f"Generic_3ph_{kva:g}kVA", "Generic", kva, 3, None, None, None,
             TRANSFORMER_LOSS_PCT, src] for kva in TRANSFORMER_SIZES_KVA]
    write_sheet(ws, ["transformer_id", "manufacturer_id", "kva_rating",
                     "phases", "primary_voltage_v", "secondary_voltage_min_v",
                     "secondary_voltage_max_v", "loss_pct", "source"], rows)
    write_readme(wb, "Catálogo de transformadores de superficie", [
        ("Claves y relaciones",
         [("transformer_id", "PK: ej. Generic_3ph_100kVA."),
          ("manufacturer_id", "FK → manufacturers. 'Generic' = tamaño estándar de norma, sin fabricante específico."),
          ("Origen de estos datos", "Hasta el esquema v2, los tamaños estándar vivían hardcodeados en core/electrical.py. Se extrajeron a esta tabla cumpliendo la regla del proyecto: los datos de ingeniería viven fuera del código.")]),
        ("Columnas",
         [("kva_rating", "Potencia aparente nominal [kVA]. Restricción: > 0. Selección: menor kVA ≥ demanda (kVA = Vs × I × √3 / 1000)."),
          ("phases", "3 = unidad trifásica; 1 = unidad monofásica (se instalan bancos de 3)."),
          ("primary_voltage_v", "Tensión primaria [V] (vacío en tamaños genéricos; completar con catálogos reales)."),
          ("secondary_voltage_min_v / max_v", "Rango del secundario con taps [V]."),
          ("loss_pct", "Pérdida típica secundario→primario [%]; el código usa 2.5% por defecto."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "transformers.xlsx")


def build_vsds():
    wb = Workbook()
    ws = wb.active
    ws.title = "vsds"
    write_sheet(ws, ["vsd_id", "manufacturer_id", "model", "kva_rating",
                     "input_voltage_v", "output_voltage_max_v",
                     "output_freq_min_hz", "output_freq_max_hz",
                     "current_max_a", "drive_type", "nema_rating", "source"],
                [])
    write_readme(wb, "Catálogo de variadores de frecuencia (VSD) — PLANTILLA", [
        ("Estado",
         [("Plantilla vacía", "La aplicación hoy solo usa el flag use_vsd (DesignObjectives). Cuando se implemente el diseño con VSD (leyes de afinidad, Brown §4.5327) y se carguen catálogos desde TESIS/CATALOGOS, esta tabla ya tiene la estructura definida.")]),
        ("Claves y relaciones",
         [("vsd_id", "PK: fabricante_modelo."),
          ("manufacturer_id", "FK → manufacturers. Relación 1:N.")]),
        ("Columnas",
         [("kva_rating", "Capacidad [kVA]. Restricción: > 0."),
          ("input_voltage_v / output_voltage_max_v", "Tensión de entrada / máxima de salida [V]."),
          ("output_freq_min_hz / max_hz", "Rango de frecuencia de salida [Hz]; típico 30–90 Hz."),
          ("current_max_a", "Corriente máxima de salida [A]."),
          ("drive_type", "Topología: 6-pulse, 12-pulse o PWM. Valores permitidos: esos tres."),
          ("nema_rating", "Tipo de gabinete (NEMA 1, NEMA 3R...)."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "vsds.xlsx")


# ======================================================================
# well_examples (desde data/example_wells.json) y real_wells (plantilla)
# ======================================================================
WELL_BLOCKS = ["reservoir", "fluid", "well", "surface", "objectives",
               "book_reference"]
BLOCK_SHEET = {"reservoir": "reservoir", "fluid": "fluid",
               "well": "well_geometry", "surface": "surface_conditions",
               "objectives": "design_objectives",
               "book_reference": "book_reference"}


def build_well_examples():
    with open(DATA_DIR / "example_wells.json", encoding="utf-8") as f:
        data = json.load(f)
    wells = {k: v for k, v in data.items() if not k.startswith("_")}

    wb = Workbook()
    ws = wb.active
    ws.title = "wells"
    write_sheet(ws, ["well_id", "description", "source"],
                [[wid, w.get("description", ""), "data/example_wells.json"]
                 for wid, w in wells.items()])

    for block in WELL_BLOCKS:
        # columnas = unión de claves ESCALARES en orden de primera aparición
        # (los valores anidados van a book_reference_details, ver abajo)
        cols: list[str] = []
        for w in wells.values():
            for k, v in w.get(block, {}).items():
                if isinstance(v, (dict, list)):
                    continue
                if k not in cols:
                    cols.append(k)
        if not cols:
            continue
        ws_b = wb.create_sheet(BLOCK_SHEET[block])
        rows = [[wid] + [w.get(block, {}).get(c) for c in cols]
                for wid, w in wells.items()]
        write_sheet(ws_b, ["well_id"] + cols, rows)

    # Detalles de validación anidados (ej. incrementos de presión del
    # Ejemplo 3A, casos del Ejemplo 3B): tabla clave-valor en 1FN.
    # PK compuesta (well_id, detail_group, item).
    detail_rows = []
    for wid, w in wells.items():
        for group, v in w.get("book_reference", {}).items():
            if isinstance(v, dict):
                for item, val in v.items():
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val)
                    detail_rows.append([wid, group, item, val])
    if detail_rows:
        ws_d = wb.create_sheet("book_reference_details")
        write_sheet(ws_d, ["well_id", "detail_group", "item", "value"],
                    detail_rows)

    write_readme(wb, "Pozos de ejemplo — casos de validación (Brown Vol. 2B)", [
        ("Estructura y claves",
         [("wells", "Tabla maestra. PK: well_id (ej. example_1a)."),
          ("Hojas 1:1", "reservoir, fluid, well_geometry, surface_conditions, design_objectives y book_reference tienen relación 1:1 con wells: misma PK well_id, que es a la vez FK → wells."),
          ("Por qué 1:1 y no una tabla ancha", "Cada hoja es el espejo de un dataclass de core/models.py (Reservoir, Fluid, WellGeometry, SurfaceConditions, DesignObjectives): misma responsabilidad, mismos campos. Autodocumentante y validable por bloque.")]),
        ("Unidades (convención del proyecto)",
         [("Presiones", "psia (diferenciales en psi)."),
          ("Temperaturas", "°F."),
          ("Caudales", "STB/d o bpd."),
          ("Profundidades", "pies (TVD o MD)."),
          ("Diámetros", "pulgadas.")]),
        ("book_reference",
         [("Contenido", "Resultados esperados del libro (TDH, etapas, HP, bomba) usados como regresión: los tests comparan el cálculo de la aplicación contra estos valores."),
          ("book_reference_details", "Tabla clave-valor para conjuntos de validación heterogéneos (incrementos de presión del Ej. 3A, casos del Ej. 3B). PK compuesta: (well_id, detail_group, item). Mantiene la 1FN sin crear decenas de columnas dispersas.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "well_examples.xlsx")
    return wells


def build_real_wells(example_wells):
    """Plantilla: misma estructura que well_examples + field_cases + fluid_samples."""
    wb = Workbook()
    ws = wb.active
    ws.title = "wells"
    write_sheet(ws, ["well_id", "field", "operator", "description", "source"], [])

    # mismas columnas que los ejemplos (sin book_reference)
    for block in ["reservoir", "fluid", "well", "surface", "objectives"]:
        cols: list[str] = []
        for w in example_wells.values():
            for k in w.get(block, {}):
                if k not in cols:
                    cols.append(k)
        ws_b = wb.create_sheet(BLOCK_SHEET[block])
        write_sheet(ws_b, ["well_id"] + cols, [])

    ws_fc = wb.create_sheet("field_cases")
    write_sheet(ws_fc, ["case_id", "well_id", "install_date", "pump_id",
                        "motor_id", "cable_id", "seal_id", "gas_handler_id",
                        "transformer_id", "vsd_id", "operating_frequency_hz",
                        "measured_flow_bpd", "measured_pip_psi", "status",
                        "failure_date", "failure_cause", "notes", "source"], [])

    ws_fs = wb.create_sheet("fluid_samples")
    write_sheet(ws_fs, ["well_id", "sample_id", "pressure_psia", "temp_f",
                        "rs_scf_stb", "bo_rb_stb", "oil_viscosity_cp",
                        "gas_z_factor", "notes", "source"], [])

    write_readme(wb, "Pozos reales y casos de campo — PLANTILLA", [
        ("Estado",
         [("Plantilla vacía", "Estructura lista para cargar los datos de TESIS/Casos reales (000100w7mb0001.xlsx, cf-194, ncf-182, etc.). Misma estructura 1:1 que well_examples.xlsx.")]),
        ("Tablas adicionales",
         [("field_cases", "PK: case_id. FK: well_id → wells, más FKs OPCIONALES al equipamiento instalado (pump_id, motor_id, cable_id, seal_id, gas_handler_id, transformer_id, vsd_id). Es la tabla que conecta los catálogos con los pozos: permite comparar diseño teórico vs comportamiento real."),
          ("fluid_samples", "PK compuesta: (well_id, sample_id). Datos PVT de laboratorio para validar las correlaciones (Standing, Beggs-Robinson) contra mediciones reales.")]),
        ("Columnas de field_cases",
         [("operating_frequency_hz", "Frecuencia de operación [Hz] (≠ 60 si hay VSD)."),
          ("measured_flow_bpd / measured_pip_psi", "Caudal [bpd] y presión de admisión [psi] medidos en campo."),
          ("status", "Valores sugeridos: running, pulled, failed, standby."),
          ("failure_date / failure_cause", "Para análisis de fallas y vida útil (run life)."),
          ("source", "Trazabilidad del dato.")]),
        KEY_NOTE,
    ])
    wb.save(OUTPUT_DIR / "real_wells.xlsx")


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    manufacturers = set()
    manufacturers |= build_pumps()
    manufacturers |= build_motors()
    manufacturers |= build_cables()
    manufacturers |= build_seals()
    manufacturers |= build_gas_handlers()
    manufacturers |= build_sensors()
    manufacturers.add("Generic")
    build_manufacturers(manufacturers)
    build_transformers()
    build_vsds()
    examples = build_well_examples()
    build_real_wells(examples)
    print(f"OK — base de datos v2 generada en {OUTPUT_DIR}: "
          f"11 archivos, {len(manufacturers)} fabricantes, "
          f"{len(examples)} pozos de ejemplo")


if __name__ == "__main__":
    main()
