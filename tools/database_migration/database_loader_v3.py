"""
Loader de la base de datos Excel — esquema v3
=============================================

``ExcelCatalogManager`` con la misma interfaz que
``catalogs.loader.CatalogManager``, leyendo el esquema v3.

Diferencias con el loader v2 (reflejan la auditoría):
* la caída de tensión se lee de ``conductor_voltage_drop`` y se asocia a
  cada cable por (conductor, size) — el JOIN de la corrección 3FN (H1);
* las columnas ``source_id`` se resuelven contra ``data_sources.xlsx``
  para reconstruir la cita completa en ``_source`` (H6) — así el resto
  de la aplicación no nota ningún cambio.

Al integrar a la aplicación, este archivo se renombrará a
``database_loader.py`` definitivo.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from bes.catalogs.loader import CatalogManager
from bes.core.models import PumpCurve, PumpPerformancePoint

_EXCEL_DIR = Path(__file__).parent / "data_excel"


def _read_sheet(path: Path, sheet: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    out = [dict(zip(headers, row)) for row in rows
           if not all(v is None for v in row)]
    wb.close()
    return out


class ExcelCatalogManager(CatalogManager):
    """Misma interfaz que CatalogManager; datos desde data_excel/ (v3)."""

    def __init__(self, excel_dir: Optional[str] = None) -> None:
        base = Path(excel_dir) if excel_dir else _EXCEL_DIR
        # citas completas: source_id -> citation (H6)
        self._citations = {
            r["source_id"]: r["citation"]
            for r in _read_sheet(base / "data_sources.xlsx", "data_sources")
        }
        self._pumps = self._load_pumps(base / "pumps.xlsx")
        self._motors = [self._entry(r, "motor_id")
                        for r in _read_sheet(base / "motors.xlsx", "motors")]
        self._cables = self._load_cables(base / "cables.xlsx")
        self._seals = self._load_seals(base / "seals.xlsx")
        self._gas_handlers = [
            self._entry(r, "gas_handler_id")
            for r in _read_sheet(base / "gas_handlers.xlsx", "gas_handlers")]
        self._sensors = [
            self._entry(r, "sensor_id")
            for r in _read_sheet(base / "sensors.xlsx", "sensors")]

    # ------------------------------------------------------------------
    def _entry(self, rec: dict, id_col: str) -> dict:
        """Fila Excel → dict con las claves que espera el código."""
        out = {}
        for k, v in rec.items():
            if k == id_col:
                continue
            if k == "manufacturer_id":
                out["manufacturer"] = v
            elif k == "source_id":
                out["_source"] = self._citations.get(v, v)
            elif k == "range_source_id":
                out["_range_source"] = self._citations.get(v, v)
            else:
                out[k] = v
        if out.get("series") is not None:
            out["series"] = str(out["series"])
        return out

    def _load_pumps(self, path: Path) -> list[PumpCurve]:
        curves: dict[str, list[PumpPerformancePoint]] = {}
        for r in _read_sheet(path, "pump_curves"):
            curves.setdefault(r["pump_id"], []).append(PumpPerformancePoint(
                flow_rate=r["flow_bpd"],
                head_per_stage=r["head_ft_per_stage"],
                hp_per_stage=r["hp_per_stage"],
                efficiency=r["efficiency"]))
        housings: dict[str, list[int]] = {}
        for r in _read_sheet(path, "pump_housings"):
            housings.setdefault(r["pump_id"], []).append(int(r["stages"]))
        pumps = []
        self.pumps_without_curves: list[str] = []
        for m in _read_sheet(path, "pumps"):
            pid = m["pump_id"]
            if pid not in curves:
                # v3.1: bombas de catálogo cuya curva aún no fue
                # digitalizada (ej. Alkhorayef: las curvas son gráficos).
                # Quedan en la base pero NO disponibles para diseño.
                self.pumps_without_curves.append(pid)
                continue
            if pid not in housings:
                raise ValueError(f"Bomba '{pid}' con curva pero sin housings")
            pumps.append(PumpCurve(
                manufacturer=m["manufacturer_id"], series=str(m["series"]),
                model=str(m["model"]), od=m["od_inches"],
                min_flow=m["min_flow_bpd"], max_flow=m["max_flow_bpd"],
                bep_flow=m["bep_flow_bpd"], max_stages=m["max_stages"],
                housing_options=sorted(housings[pid]),
                points=sorted(curves[pid], key=lambda p: p.flow_rate)))
        return pumps

    def _load_cables(self, path: Path) -> list[dict]:
        # JOIN v3: caída de tensión por (conductor, size) — H1
        vdrops: dict[tuple, dict[str, float]] = {}
        for r in _read_sheet(path, "conductor_voltage_drop"):
            key = (r["conductor"], str(r["size"]))
            vdrops.setdefault(key, {})[str(int(r["temp_f"]))] = (
                r["v_per_amp_per_1000ft"])
        cables = []
        for m in _read_sheet(path, "cables"):
            c = self._entry(m, "cable_id")
            c["size"] = str(c["size"])
            key = (c["conductor"], c["size"])
            if key not in vdrops:
                raise ValueError(
                    f"Sin caída de tensión para conductor {key} "
                    f"en 'conductor_voltage_drop'")
            c["voltage_drop_v_per_amp_per_1000ft"] = dict(vdrops[key])
            cables.append(c)
        return cables

    def _load_seals(self, path: Path) -> list[dict]:
        compat: dict[str, list[str]] = {}
        for r in _read_sheet(path, "seal_motor_compatibility"):
            compat.setdefault(r["seal_id"], []).append(str(r["motor_series"]))
        seals = []
        for m in _read_sheet(path, "seals"):
            sid = m["seal_id"]
            s = self._entry(m, "seal_id")
            s["compatible_motor_series"] = compat.get(sid, [])
            seals.append(s)
        return seals

    # ------------------------------------------------------------------
    def get_transformer_sizes_kva(self) -> list[float]:
        rows = _read_sheet(_EXCEL_DIR / "transformers.xlsx", "transformers")
        return sorted(float(r["kva_rating"]) for r in rows)
