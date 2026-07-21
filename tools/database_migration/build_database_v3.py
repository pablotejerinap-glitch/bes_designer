"""
Constructor de la base de datos Excel — esquema v3 (post-auditoría)
===================================================================

Aplica los hallazgos de AUDITORIA_BASE_DE_DATOS.md sobre el esquema v2:

* H1: ``conductor_voltage_drop`` (PK conductor+size+temp_f) reemplaza a
  ``cable_voltage_drop`` — corrige la violación de 3FN (la caída de
  tensión es física del conductor/calibre, no del producto comercial).
* H2: nueva tabla ``equipment_series`` — las series (375, 400, 513...)
  pasan a ser entidad con FK real; documenta que la serie 513 existe en
  bombas pero aún no tiene motores cargados.
* H3: ``installation_components`` reemplaza las FKs simples de
  ``field_cases`` — permite ensambles reales en tándem.
* H4: columna discriminadora ``well_type`` en las tablas wells
  (example/real); en SQLite serán UNA tabla.
* H5: nueva plantilla ``switchboards`` (tableros de control).
* H6: nueva tabla ``data_sources`` — cada fuente bibliográfica recibe un
  ID corto (BROWN-01, CHX-02...); las columnas ``source`` de todas las
  tablas pasan a contener ese source_id.

Supersede a build_database.py (v2). No modifica el proyecto original.

USO:  python build_database_v3.py
"""
from __future__ import annotations

import json
from pathlib import Path

import bes

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
# tools/database_migration/ -> raiz del repo
_REPO_ROOT = BASE_DIR.parent.parent
# Los catalogos JSON son datos del paquete instalable: se resuelven a
# traves de el, no por ruta relativa al repo.
_BES_PKG = Path(bes.__file__).parent
CATALOG_DIR = _BES_PKG / "catalogs"
DATA_DIR = _REPO_ROOT / "backend" / "data"
OUTPUT_DIR = BASE_DIR / "data_excel"

FONT = "Arial"
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
CELL_FONT = Font(name=FONT)

TRANSFORMER_SIZES_KVA = (25.0, 37.5, 50.0, 75.0, 100.0, 150.0, 200.0, 300.0)
TRANSFORMER_LOSS_PCT = 2.5

MANUFACTURER_INFO = {
    "Reda": ("Schlumberger REDA (SLB)", "EE.UU."),
    "Centrilift": ("Baker Hughes Centrilift", "EE.UU."),
    "SLB": ("SLB (Schlumberger)", "EE.UU."),
    "Weatherford": ("Weatherford International", "EE.UU."),
    "ChampionX": ("ChampionX Artificial Lift", "EE.UU."),
    "Generic": ("Tamaños estándar de norma (sin fabricante)", ""),
}

# ----------------------------------------------------------------------
# Registro de fuentes (H6): cadena completa -> source_id corto
# ----------------------------------------------------------------------
_SOURCES: dict[str, str] = {}          # citation -> source_id
_PREFIX_COUNT: dict[str, int] = {}


def _prefix(s: str) -> str:
    sl = s.lower()
    if "brown" in sl:
        return "BROWN"
    if "championx" in sl:
        return "CHX"
    if "slb" in sl or "affirmed" in sl:
        return "SLB"
    if "reda" in sl:
        return "REDA"
    if "centrilift" in sl:
        return "CENT"
    if sl.startswith("ace"):
        return "ACE"
    if "weatherford" in sl:
        return "WFT"
    if "api rp" in sl:
        return "API"
    return "SRC"


def src_id(citation: str | None) -> str | None:
    """Registra una fuente y devuelve su ID corto (idempotente)."""
    if not citation:
        return None
    citation = citation.strip()
    if citation not in _SOURCES:
        p = _prefix(citation)
        _PREFIX_COUNT[p] = _PREFIX_COUNT.get(p, 0) + 1
        _SOURCES[citation] = f"{p}-{_PREFIX_COUNT[p]:02d}"
    return _SOURCES[citation]


def _source_type(citation: str) -> str:
    sl = citation.lower()
    if "estimated" in sl or "no confirmada" in sl or "aprox" in sl:
        return "estimado / por confirmar"
    if "brown" in sl and "vol" in sl:
        return "libro"
    if "data sheet" in sl or "catalog" in sl or "catálogo" in sl:
        return "catálogo"
    if "api rp" in sl or "nema" in sl or "ansi" in sl or "ieee" in sl:
        return "norma"
    return "nota"


# ----------------------------------------------------------------------
# Helpers de escritura (idénticos a v2)
# ----------------------------------------------------------------------
def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = CELL_FONT
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(width + 3, 60)
    ws.freeze_panes = "A2"


def write_readme(wb, title, sections):
    ws = wb.create_sheet("README")
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    r = 3
    for heading, items in sections:
        ws[f"A{r}"] = heading
        ws[f"A{r}"].font = Font(name=FONT, bold=True, size=12, color="1F4E79")
        r += 1
        for item, desc in items:
            ws[f"A{r}"] = item
            ws[f"B{r}"] = desc
            ws[f"A{r}"].font = Font(name=FONT, bold=True)
            ws[f"B{r}"].font = CELL_FONT
            ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 105


def load_json(directory, name, key):
    with open(directory / name, encoding="utf-8") as f:
        return json.load(f)[key]


def make_id(e):
    parts = [e["manufacturer"], e.get("series", ""), e["model"]]
    return "_".join(str(p) for p in parts if p)


def cable_id(c):
    return (f"{c['manufacturer']}_{c['type']}_{c['conductor']}"
            f"_{c['size']}").replace(" ", "")


KEY_NOTE = ("Convención de claves",
            [("PK", "Clave primaria (PRIMARY KEY en SQLite)."),
             ("FK", "Clave foránea (FOREIGN KEY en SQLite)."),
             ("source_id", "FK → data_sources.xlsx: toda fila de ingeniería referencia su fuente."),
             ("Celda vacía", "NULL: dato no aplicable o no disponible. Nunca 0 como 'sin dato'.")])


# ======================================================================
def build_pumps():
    pumps = load_json(CATALOG_DIR, "pumps.json", "pumps")
    wb = Workbook()
    ws = wb.active
    ws.title = "pumps"
    write_sheet(ws, ["pump_id", "manufacturer_id", "series", "model",
                     "od_inches", "min_flow_bpd", "max_flow_bpd",
                     "bep_flow_bpd", "max_stages", "source_id"],
                [[make_id(p), p["manufacturer"], p["series"], p["model"],
                  p["od_inches"], p["min_flow_bpd"], p["max_flow_bpd"],
                  p["bep_flow_bpd"], p["max_stages"],
                  src_id(p.get("_source"))] for p in pumps])
    ws2 = wb.create_sheet("pump_curves")
    write_sheet(ws2, ["pump_id", "flow_bpd", "head_ft_per_stage",
                      "hp_per_stage", "efficiency"],
                [[make_id(p), pt["flow_bpd"], pt["head_ft_per_stage"],
                  pt["hp_per_stage"], pt["efficiency"]]
                 for p in pumps for pt in p["performance_curve"]])
    ws3 = wb.create_sheet("pump_housings")
    write_sheet(ws3, ["pump_id", "stages"],
                [[make_id(p), h] for p in pumps for h in p["housing_options"]])
    write_readme(wb, "Catálogo de bombas electrosumergibles (BES/ESP)", [
        ("Tablas, claves y relaciones",
         [("pumps", "Maestra. PK: pump_id. FK: manufacturer_id → manufacturers; series → equipment_series; source_id → data_sources."),
          ("pump_curves", "PK compuesta (pump_id, flow_bpd). FK pump_id → pumps. 1:N. Curvas a 60 Hz con agua (SG=1.0)."),
          ("pump_housings", "PK compuesta (pump_id, stages). FK pump_id → pumps. 1:N.")]),
        ("Columnas",
         [("min/max/bep_flow_bpd", "Valores DECLARADOS por el fabricante (rango recomendado y BEP de catálogo); no se derivan de la curva cargada — pueden no coincidir con el máximo discreto de eficiencia (hallazgo H7 de la auditoría)."),
          ("head_ft_per_stage / hp_per_stage", "[pies/etapa] y [HP/etapa] con agua; hp se multiplica por SG del fluido real."),
          ("efficiency", "[0–1]. Restricción: 0 < eff < 1 (la verifica check_integrity.py)."),
          ("max_stages", "Entero > 0.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "pumps.xlsx")
    return pumps


def build_motors():
    motors = load_json(CATALOG_DIR, "motors.json", "motors")
    wb = Workbook()
    ws = wb.active
    ws.title = "motors"
    write_sheet(ws, ["motor_id", "manufacturer_id", "series", "model",
                     "hp_rating", "voltage", "amperage", "length_ft",
                     "max_temp_f", "od_inches", "source_id"],
                [[make_id(m), m["manufacturer"], m["series"], m["model"],
                  m["hp_rating"], m["voltage"], m["amperage"],
                  m["length_ft"], m["max_temp_f"], m["od_inches"],
                  src_id(m.get("_source"))] for m in motors])
    write_readme(wb, "Catálogo de motores electrosumergibles", [
        ("Claves y relaciones",
         [("motor_id", "PK: fabricante_serie_modelo."),
          ("manufacturer_id / series / source_id", "FK → manufacturers / equipment_series / data_sources."),
          ("Decisión H9 (auditoría)", "No se divide en frame + devanados: el catálogo actual no tiene modelos con múltiples variantes V/A. Revisar al cargar catálogos que las publiquen.")]),
        ("Columnas",
         [("hp_rating", "[HP] a 60 Hz. > 0."),
          ("voltage / amperage", "Placa [V] / [A]."),
          ("max_temp_f", "[°F], comparar contra temperatura de fondo.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "motors.xlsx")
    return motors


def build_cables():
    """H1: caída de tensión normalizada a (conductor, size, temp_f)."""
    cables = load_json(CATALOG_DIR, "cables.json", "cables")
    wb = Workbook()
    ws = wb.active
    ws.title = "cables"
    write_sheet(ws, ["cable_id", "manufacturer_id", "type", "conductor",
                     "size", "max_amps", "max_temp_f", "source_id"],
                [[cable_id(c), c["manufacturer"], c["type"], c["conductor"],
                  c["size"], c["max_amps"], c["max_temp_f"],
                  src_id(c.get("_source"))] for c in cables])

    seen: dict[tuple, dict] = {}
    for c in cables:
        key = (c["conductor"], c["size"])
        if key not in seen:
            seen[key] = c["voltage_drop_v_per_amp_per_1000ft"]
    rows = []
    for (cond, size), vd in sorted(seen.items()):
        for t, v in sorted(vd.items(), key=lambda kv: float(kv[0])):
            rows.append([cond, size, int(t), v])
    ws2 = wb.create_sheet("conductor_voltage_drop")
    write_sheet(ws2, ["conductor", "size", "temp_f",
                      "v_per_amp_per_1000ft"], rows)

    write_readme(wb, "Catálogo de cables de potencia para BES", [
        ("Tablas, claves y relaciones",
         [("cables", "Maestra (producto comercial: tipo, aislación, temperatura). PK: cable_id (incluye conductor CU/AL). FK: manufacturer_id → manufacturers; (conductor, size) → conductor_voltage_drop; source_id → data_sources."),
          ("conductor_voltage_drop", "Propiedad FÍSICA del conductor. PK compuesta (conductor, size, temp_f)."),
          ("Hallazgo H1 de la auditoría", "En el esquema v2 la caída colgaba de cable_id, duplicando los mismos valores en cada producto que compartía conductor y calibre (violación de 3FN verificada: 4 combinaciones compartidas, 0 con curvas distintas). En v3 cada par (conductor, calibre) tiene UNA curva; corregir un valor se hace en un solo lugar.")]),
        ("Columnas",
         [("conductor", "CU (cobre) o AL (aluminio)."),
          ("size", "Calibre AWG."),
          ("max_amps", "Ampacidad [A] del producto (valor único conservador; el derating por temperatura es una tabla futura: cable_ampacity_derating)."),
          ("v_per_amp_per_1000ft", "[V/A/1000 pies]; el código interpola linealmente en temperatura.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "cables.xlsx")
    return cables


def build_seals():
    seals = load_json(CATALOG_DIR, "seals.json", "seals")
    wb = Workbook()
    ws = wb.active
    ws.title = "seals"
    write_sheet(ws, ["seal_id", "manufacturer_id", "series", "model", "type",
                     "od_inches", "length_ft", "thrust_capacity_lbs",
                     "max_temp_f", "shaft_hp_standard",
                     "shaft_hp_high_strength", "source_id"],
                [[make_id(s), s["manufacturer"], s["series"], s["model"],
                  s["type"], s["od_inches"], s["length_ft"],
                  s["thrust_capacity_lbs"], s["max_temp_f"],
                  s.get("shaft_hp_standard"), s.get("shaft_hp_high_strength"),
                  src_id(s.get("_source"))] for s in seals])
    ws2 = wb.create_sheet("seal_motor_compatibility")
    write_sheet(ws2, ["seal_id", "motor_series"],
                [[make_id(s), serie] for s in seals
                 for serie in s.get("compatible_motor_series", [])])
    write_readme(wb, "Catálogo de sellos / protectores", [
        ("Tablas, claves y relaciones",
         [("seals", "Maestra. PK: seal_id. FK: manufacturer_id → manufacturers; series → equipment_series; source_id → data_sources."),
          ("seal_motor_compatibility", "Tabla puente N:M sellos ↔ series de motor. PK compuesta (seal_id, motor_series). FK: seal_id → seals; motor_series → equipment_series (FK REAL desde v3 — hallazgo H2: la serie 513 citada aquí no tenía destino verificable en v2).")]),
        ("Columnas",
         [("type", "labyrinth / bag / combined (Brown §4.5325)."),
          ("thrust_capacity_lbs", "Capacidad de empuje axial [lbs]. > 0."),
          ("shaft_hp_*", "Potencia transmisible por eje estándar / alta resistencia [HP]. Vacío = no informado.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "seals.xlsx")
    return seals


def build_gas_handlers():
    ghs = load_json(CATALOG_DIR, "gas_handlers.json", "gas_handlers")
    wb = Workbook()
    ws = wb.active
    ws.title = "gas_handlers"
    write_sheet(ws, ["gas_handler_id", "manufacturer_id", "series", "model",
                     "type", "position", "od_inches", "length_ft",
                     "weight_lbs", "hp", "min_flow_bpd", "max_flow_bpd",
                     "max_efficiency", "source_id", "range_source_id"],
                [[make_id(g), g["manufacturer"], g["series"], g["model"],
                  g["type"], g["position"], g["od_inches"], g["length_ft"],
                  g["weight_lbs"], g["hp"], g["min_flow_bpd"],
                  g["max_flow_bpd"], g["max_efficiency"],
                  src_id(g.get("_source")), src_id(g.get("_range_source"))]
                 for g in ghs])
    write_readme(wb, "Catálogo de separadores / manejadores de gas", [
        ("Claves y relaciones",
         [("gas_handler_id", "PK. FK: manufacturer_id → manufacturers; source_id y range_source_id → data_sources (el equipo y su rango operativo pueden tener fuentes distintas).")]),
        ("Columnas",
         [("type", "vortex / rotary / etc. Los vortex logran la mayor eficiencia."),
          ("min/max_flow_bpd", "Rango operativo [bpd], min < max."),
          ("max_efficiency", "Eficiencia máxima de separación [0–1].")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "gas_handlers.xlsx")
    return ghs


def build_sensors():
    sensors = load_json(CATALOG_DIR, "sensors.json", "sensors")
    wb = Workbook()
    ws = wb.active
    ws.title = "sensors"
    write_sheet(ws, ["sensor_id", "manufacturer_id", "model",
                     "intake_pressure_max_psi", "intake_temp_max_f",
                     "discharge_pressure_max_psi", "motor_winding_temp_max_f",
                     "vibration_monitoring", "vibration_max_g", "od_inches",
                     "length_in", "weight_lbs", "max_motor_voltage",
                     "source_id"],
                [[f"{s['manufacturer']}_{s['model']}".replace(" ", ""),
                  s["manufacturer"], s["model"], s["intake_pressure_max_psi"],
                  s["intake_temp_max_f"], s.get("discharge_pressure_max_psi"),
                  s["motor_winding_temp_max_f"], s["vibration_monitoring"],
                  s.get("vibration_max_g"), s["od_inches"], s["length_in"],
                  s["weight_lbs"], s["max_motor_voltage"],
                  src_id(s.get("_source"))] for s in sensors])
    write_readme(wb, "Catálogo de sensores de fondo", [
        ("Claves y relaciones",
         [("sensor_id", "PK: fabricante_modelo. FK: manufacturer_id → manufacturers; source_id → data_sources.")]),
        ("Columnas",
         [("discharge_pressure_max_psi", "Vacío (NULL) = no mide descarga."),
          ("vibration_monitoring / vibration_max_g", "TRUE/FALSE y rango [g]; vacío = no aplica.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "sensors.xlsx")
    return sensors


def build_transformers():
    wb = Workbook()
    ws = wb.active
    ws.title = "transformers"
    sid = src_id("Tamaños estándar trifásicos ANSI/IEEE; extraído de "
                 "core/electrical.py::_TRANSFORMER_SIZES_KVA")
    rows = [[f"Generic_3ph_{kva:g}kVA", "Generic", kva, 3, None, None, None,
             TRANSFORMER_LOSS_PCT, sid] for kva in TRANSFORMER_SIZES_KVA]
    write_sheet(ws, ["transformer_id", "manufacturer_id", "kva_rating",
                     "phases", "primary_voltage_v", "secondary_voltage_min_v",
                     "secondary_voltage_max_v", "loss_pct", "source_id"], rows)
    write_readme(wb, "Catálogo de transformadores de superficie", [
        ("Claves y relaciones",
         [("transformer_id", "PK. FK: manufacturer_id → manufacturers; source_id → data_sources."),
          ("Origen", "Datos extraídos de core/electrical.py (regla del proyecto: los datos de ingeniería viven fuera del código).")]),
        ("Columnas",
         [("kva_rating", "[kVA] > 0. Selección: menor kVA ≥ demanda (kVA = Vs × I × √3 / 1000)."),
          ("phases", "3 = trifásico; 1 = monofásico (bancos de 3)."),
          ("loss_pct", "Pérdida típica [%]; el código usa 2.5% por defecto.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "transformers.xlsx")


def build_vsds():
    wb = Workbook()
    ws = wb.active
    ws.title = "vsds"
    write_sheet(ws, ["vsd_id", "manufacturer_id", "model", "kva_rating",
                     "input_voltage_v", "output_voltage_max_v",
                     "output_freq_min_hz", "output_freq_max_hz",
                     "current_max_a", "drive_type", "nema_rating",
                     "source_id"], [])
    write_readme(wb, "Catálogo de variadores de frecuencia (VSD) — PLANTILLA", [
        ("Estado", [("Plantilla vacía", "Estructura lista para los catálogos de TESIS/CATALOGOS y el diseño con leyes de afinidad (Brown §4.5327).")]),
        ("Claves", [("vsd_id", "PK. FK: manufacturer_id → manufacturers; source_id → data_sources.")]),
        ("Columnas",
         [("output_freq_min/max_hz", "Rango de frecuencia [Hz]; típico 30–90."),
          ("drive_type", "6-pulse / 12-pulse / PWM (valores permitidos).")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "vsds.xlsx")


def build_switchboards():
    """H5: tableros de control (catálogo en TESIS/CATALOGOS/Swbd-Mtr Cntrl.pdf)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "switchboards"
    write_sheet(ws, ["switchboard_id", "manufacturer_id", "model",
                     "max_voltage_v", "max_amps", "max_hp", "nema_rating",
                     "source_id"], [])
    write_readme(wb, "Catálogo de tableros de control (switchboards) — PLANTILLA", [
        ("Estado", [("Plantilla vacía", "Hallazgo H5 de la auditoría: todo pozo BES sin VSD se controla con switchboard; el catálogo ya está en TESIS/CATALOGOS (Swbd-Mtr Cntrl.pdf).")]),
        ("Claves", [("switchboard_id", "PK: fabricante_modelo. FK: manufacturer_id → manufacturers; source_id → data_sources.")]),
        ("Columnas",
         [("max_voltage_v / max_amps / max_hp", "Límites eléctricos del tablero."),
          ("nema_rating", "Tipo de gabinete (NEMA 1, 3R...).")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "switchboards.xlsx")


def build_equipment_series(pumps, motors, seals):
    """H2: las series pasan a ser entidad con FK real."""
    usage: dict[str, set] = {}
    for p in pumps:
        usage.setdefault(str(p["series"]), set()).add("pumps")
    for m in motors:
        usage.setdefault(str(m["series"]), set()).add("motors")
    for s in seals:
        usage.setdefault(str(s["series"]), set()).add("seals")
        for ms in s.get("compatible_motor_series", []):
            usage.setdefault(str(ms), set()).add("seal_motor_compatibility")
    rows = []
    for serie in sorted(usage, key=lambda x: (not x.isdigit(), x)):
        od = round(int(serie) / 100.0, 2) if serie.isdigit() else None
        used_in = ", ".join(sorted(usage[serie]))
        note = ""
        if "seal_motor_compatibility" in usage[serie] and "motors" not in usage[serie]:
            note = ("Serie citada como compatible por sellos pero SIN motores "
                    "cargados aún en el catálogo (hallazgo H2).")
        rows.append([serie, od, used_in, note])
    wb = Workbook()
    ws = wb.active
    ws.title = "equipment_series"
    write_sheet(ws, ["series_id", "od_nominal_inches", "used_in", "notes"],
                rows)
    write_readme(wb, "Series de equipos (clases de diámetro)", [
        ("Qué es una serie",
         [("Concepto", "La serie identifica la clase de diámetro del equipo: por convención de la industria, el número ≈ OD en centésimas de pulgada (serie 400 ≈ 4.00\", serie 540 ≈ 5.40\"). Determina en qué casing entra el equipo y qué componentes son acoplables entre sí.")]),
        ("Claves y relaciones",
         [("series_id", "PK. La referencian como FK: pumps.series, motors.series, seals.series y seal_motor_compatibility.motor_series."),
          ("Hallazgo H2", "Esta tabla existe porque en v2 'seal_motor_compatibility.motor_series' no tenía destino, y los datos contenían la serie 513 sin motores: un error invisible para el modelo. Ahora es una fila documentada y verificable.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "equipment_series.xlsx")


def build_manufacturers(all_m):
    wb = Workbook()
    ws = wb.active
    ws.title = "manufacturers"
    rows = []
    for m in sorted(all_m):
        full, country = MANUFACTURER_INFO.get(m, (m, ""))
        rows.append([m, full, country, ""])
    write_sheet(ws, ["manufacturer_id", "full_name", "country", "notes"], rows)
    write_readme(wb, "Tabla maestra de fabricantes", [
        ("Claves y relaciones",
         [("manufacturer_id", "PK. 1:N con todos los catálogos de equipos (incl. transformers, vsds y switchboards)."),
          ("Cómo agregar un fabricante", "Una fila aquí + sus equipos en los otros archivos. Cero código.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "manufacturers.xlsx")


def build_data_sources():
    """H6: bibliografía normalizada. Se llama AL FINAL (registro completo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data_sources"
    rows = [[sid, _source_type(cit), cit]
            for cit, sid in sorted(_SOURCES.items(), key=lambda kv: kv[1])]
    write_sheet(ws, ["source_id", "type", "citation"], rows)
    write_readme(wb, "Registro de fuentes de datos (bibliografía)", [
        ("Claves y relaciones",
         [("source_id", "PK corta y legible (BROWN-01, CHX-02...). La referencian TODAS las columnas source_id de la base."),
          ("Hallazgo H6", "En v2 cada fila repetía la cita completa como texto libre, con riesgo de variantes de tipeo y sin poder consultar 'todo lo que salió de Brown'. Ahora la cita vive una sola vez.")]),
        ("Columnas",
         [("type", "libro / catálogo / norma / estimado por confirmar / nota. Las fuentes 'estimado' señalan datos a validar con catálogos reales de TESIS."),
          ("citation", "Cita completa original (se preserva textualmente para trazabilidad).")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "data_sources.xlsx")


WELL_BLOCKS = ["reservoir", "fluid", "well", "surface", "objectives",
               "book_reference"]
BLOCK_SHEET = {"reservoir": "reservoir", "fluid": "fluid",
               "well": "well_geometry", "surface": "surface_conditions",
               "objectives": "design_objectives",
               "book_reference": "book_reference"}


def build_well_examples():
    with open(DATA_DIR / "example_wells.json", encoding="utf-8") as f:
        data = json.load(f)
    wells = {k: v for k, v in data.items() if not k.startswith("_")}
    wb = Workbook()
    ws = wb.active
    ws.title = "wells"
    sid = src_id("data/example_wells.json — casos de validación Brown Vol. 2B")
    write_sheet(ws, ["well_id", "well_type", "description", "source_id"],
                [[wid, "example", w.get("description", ""), sid]
                 for wid, w in wells.items()])
    for block in WELL_BLOCKS:
        cols = []
        for w in wells.values():
            for k, v in w.get(block, {}).items():
                if isinstance(v, (dict, list)):
                    continue
                if k not in cols:
                    cols.append(k)
        if not cols:
            continue
        ws_b = wb.create_sheet(BLOCK_SHEET[block])
        rows = [[wid] + [w.get(block, {}).get(c) for c in cols]
                for wid, w in wells.items()]
        write_sheet(ws_b, ["well_id"] + cols, rows)
    detail_rows = []
    for wid, w in wells.items():
        for group, v in w.get("book_reference", {}).items():
            if isinstance(v, dict):
                for item, val in v.items():
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val)
                    detail_rows.append([wid, group, item, val])
    if detail_rows:
        ws_d = wb.create_sheet("book_reference_details")
        write_sheet(ws_d, ["well_id", "detail_group", "item", "value"],
                    detail_rows)
    write_readme(wb, "Pozos de ejemplo — casos de validación (Brown Vol. 2B)", [
        ("Estructura y claves",
         [("wells", "Maestra. PK: well_id. well_type='example' (hallazgo H4: entidad wells UNIFICADA con los pozos reales; en SQLite será una sola tabla con este discriminador)."),
          ("Hojas 1:1", "reservoir, fluid, well_geometry, surface_conditions, design_objectives, book_reference: PK/FK well_id, espejo de los dataclasses de core/models.py."),
          ("book_reference_details", "Clave-valor para conjuntos de validación heterogéneos. PK (well_id, detail_group, item). Decisión H8: EAV consciente, solo lectura.")]),
        ("Unidades", [("Convención", "psia, °F, bpd/STB/d, pies, pulgadas — en el nombre de cada columna.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "well_examples.xlsx")
    return wells


def build_real_wells(example_wells):
    """H3+H4: wells con well_type; ensambles via installation_components."""
    wb = Workbook()
    ws = wb.active
    ws.title = "wells"
    write_sheet(ws, ["well_id", "well_type", "field", "operator",
                     "description", "source_id"], [])
    for block in ["reservoir", "fluid", "well", "surface", "objectives"]:
        cols = []
        for w in example_wells.values():
            for k, v in w.get(block, {}).items():
                if isinstance(v, (dict, list)):
                    continue
                if k not in cols:
                    cols.append(k)
        ws_b = wb.create_sheet(BLOCK_SHEET[block])
        write_sheet(ws_b, ["well_id"] + cols, [])
    ws_fc = wb.create_sheet("field_cases")
    write_sheet(ws_fc, ["case_id", "well_id", "install_date", "pull_date",
                        "operating_frequency_hz", "measured_flow_bpd",
                        "measured_pip_psi", "status", "failure_date",
                        "failure_cause", "notes", "source_id"], [])
    ws_ic = wb.create_sheet("installation_components")
    write_sheet(ws_ic, ["case_id", "position", "equipment_type",
                        "equipment_id", "quantity", "notes"], [])
    ws_fs = wb.create_sheet("fluid_samples")
    write_sheet(ws_fs, ["well_id", "sample_id", "pressure_psia", "temp_f",
                        "rs_scf_stb", "bo_rb_stb", "oil_viscosity_cp",
                        "gas_z_factor", "notes", "source_id"], [])
    write_readme(wb, "Pozos reales y casos de campo — PLANTILLA", [
        ("Estructura y claves",
         [("wells", "PK: well_id. well_type='real' (H4: misma entidad que los ejemplos; una sola tabla en SQLite)."),
          ("field_cases", "El EVENTO de instalación. PK: case_id. FK: well_id → wells; source_id → data_sources. Fechas, frecuencia, mediciones, estado y falla."),
          ("installation_components", "El ENSAMBLE instalado (hallazgo H3): PK (case_id, position). position = orden en el string de abajo hacia arriba (1=sensor, 2=motor, 3=sello...). equipment_type ∈ {sensor, motor, seal, gas_handler, pump, cable, transformer, vsd, switchboard}; equipment_id = PK en el catálogo correspondiente. Permite tándems (2-3 bombas, sellos dobles, cable + MLE) que las FKs simples de v2 no podían representar."),
          ("fluid_samples", "PK (well_id, sample_id). Datos PVT de laboratorio para validar correlaciones.")]),
        ("Columnas de field_cases",
         [("status", "running / pulled / failed / standby."),
          ("operating_frequency_hz", "≠ 60 si hay VSD."),
          ("failure_date / failure_cause", "Para análisis de run life.")]),
        KEY_NOTE])
    wb.save(OUTPUT_DIR / "real_wells.xlsx")


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    # borra la tabla v2 reemplazada por el nuevo diseño de cables
    pumps = build_pumps()
    motors = build_motors()
    cables = build_cables()
    seals = build_seals()
    ghs = build_gas_handlers()
    sensors = build_sensors()
    build_transformers()
    build_vsds()
    build_switchboards()
    build_equipment_series(pumps, motors, seals)
    manufacturers = ({p["manufacturer"] for p in pumps}
                     | {m["manufacturer"] for m in motors}
                     | {c["manufacturer"] for c in cables}
                     | {s["manufacturer"] for s in seals}
                     | {g["manufacturer"] for g in ghs}
                     | {s["manufacturer"] for s in sensors}
                     | {"Generic"})
    build_manufacturers(manufacturers)
    examples = build_well_examples()
    build_real_wells(examples)
    build_data_sources()   # al final: el registro de fuentes ya está completo
    print(f"OK — esquema v3: 14 archivos, {len(manufacturers)} fabricantes, "
          f"{len(_SOURCES)} fuentes registradas, {len(examples)} pozos de ejemplo")


if __name__ == "__main__":
    main()
