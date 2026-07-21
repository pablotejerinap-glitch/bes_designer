"""
Importador del catálogo Alkhorayef 2019 → base de datos Excel (v3.1)
====================================================================

Primera importación de un fabricante REAL desde su catálogo PDF.
Lee ``TESIS/CATALOGOS/alkhorayef-esp-catalog-2019.pdf`` (texto extraíble)
y agrega a la base:

* fabricante ``Alkhorayef`` en manufacturers.xlsx,
* fuente ``ALKH-01`` en data_sources.xlsx,
* las series nuevas en equipment_series.xlsx,
* ~37 bombas SPECTRUM (WA…WQ) en pumps.xlsx con los CAMPOS NUEVOS v3.1
  que pidió la metodología de la cátedra:
  min_casing_size_in, housing_pressure_limit_psi,
  housing_pressure_limit_high_psi, shaft_diameter_in,
  shaft_hp_limit_std / _hs / _uhs (límites de eje @60Hz),
  stage_type, min/max_flow_floater_bpd,
* la tabla de housings real de cada bomba (housing_code, stages,
  length_ft, weight_lbs) en pump_housings.

IMPORTANTE — curvas: el catálogo publica las curvas como GRÁFICOS, no
como tablas. Estas bombas quedan en la base SIN filas en pump_curves y
por lo tanto NO disponibles para diseño (check_integrity las reporta
como aviso; el loader las salta). Cuando se digitalicen las curvas,
solo hay que agregar filas a pump_curves.

El importador es IDEMPOTENTE: borra las filas Alkhorayef existentes y
las reescribe, así puede correrse las veces que haga falta. Orden de
regeneración completa de la base:

    python build_database_v3.py      # bootstrap desde los JSON legados
    python import_alkhorayef.py      # + fabricante Alkhorayef
    python check_integrity.py        # verificación

USO:  python import_alkhorayef.py [ruta_al_pdf]
"""
from __future__ import annotations

import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data_excel"
DEFAULT_PDF = (BASE_DIR.parent.parent / "OneDrive" / "Desktop" / "TESIS"
               / "CATALOGOS" / "alkhorayef-esp-catalog-2019.pdf")
# En el entorno de trabajo la ruta puede diferir; se acepta por argumento.

MANUFACTURER = "Alkhorayef"
SOURCE_ID = "ALKH-01"
CITATION = ("Alkhorayef Petroleum Company — International ESP Systems "
            "and Solutions, Product Catalog 2019 "
            "(TESIS/CATALOGOS/alkhorayef-esp-catalog-2019.pdf)")

CELL_FONT = Font(name="Arial")

# Columnas nuevas del esquema v3.1 (se agregan si no existen)
NEW_PUMP_COLS = [
    "min_flow_floater_bpd", "max_flow_floater_bpd", "min_casing_size_in",
    "housing_pressure_limit_psi", "housing_pressure_limit_high_psi",
    "shaft_diameter_in", "shaft_hp_limit_std", "shaft_hp_limit_hs",
    "shaft_hp_limit_uhs", "stage_type",
]
NEW_HOUSING_COLS = ["housing_code", "stages_compression",
                    "stages_ar_floater", "stages_ar_compression",
                    "length_ft", "weight_lbs"]


# ----------------------------------------------------------------------
# 1. Extracción del PDF
# ----------------------------------------------------------------------
def pdf_text(pdf: Path) -> str:
    return subprocess.run(
        ["pdftotext", "-layout", str(pdf), "-"],
        capture_output=True, text=True, check=True).stdout


def _num(s: str | None) -> float | None:
    if s is None:
        return None
    return float(s.replace(",", ""))


FIELD_PATTERNS = {
    # BPD puede caer en la línea siguiente cuando el valor es ancho
    "opt_range": r"Optimum operating range\s+([\d,]+)\s+to\s+([\d,]+)",
    "floater_range": r"Floater operating range\s+([\d,]+)\s+to\s+([\d,]+)",
    "bep": r"BEP flow rate\s+([\d,]+)",
    "housing_od": r"Housing diameter\s+([\d.]+)\s*In",
    "min_casing": r"Minimum casing size\s+([\d.]+)\s*In",
    "p_limit": r"(?<!pressure )Housing pressure limit\s+([\d,]+)\s*PSI",
    "p_limit_high": r"High pressure housing limit\s+([\d,]+)\s*PSI",
    "shaft_d": r"Shaft diameter\s+([\d.]+)\s*In",
    "shaft_std": r"Standard shaft limit @ 60Hz\s+([\d,]+)\s*HP",
    "shaft_hs": r"High strength shaft limit @ 60Hz\s+([\d,]+)\s*HP",
    "shaft_uhs": r"Ultra high strength shaft limit @ 60Hz\s+([\d,]+)\s*HP",
    "stage_type": r"Stage type\s+(\w+)",
}

# el encabezado de modelo puede venir con o sin sangría (saltos de página)
MODEL_RE = re.compile(r"^\s*(W[A-Z]-\d+)\s*$", re.M)
SERIES_RE = re.compile(r"^\s*(\d{3,4})\s+series\s*$", re.M | re.I)
# Filas de la tabla de housings. Dos formatos en el catálogo:
#   6 columnas: housing | stages(floater) | ft | m | lbs | kg
#   9 columnas: housing | std floater | std compression | AR floater |
#               AR compression | ft | m | lbs | kg   (AR = abrasion resistant)
NUM_TOKEN_RE = re.compile(r"[\d][\d,\.]*")


def parse_housing_rows(chunk: str) -> list[dict]:
    rows = []
    for line in chunk.splitlines():
        toks = NUM_TOKEN_RE.findall(line)
        # 6 = housing simple (floater) | 7 = solo compression (+AR) |
        # 9 = floater + compression + AR floater + AR compression
        if len(toks) not in (6, 7, 9):
            continue
        if line.strip() and not re.fullmatch(r"[\s\d,\.]+", line):
            continue  # la línea tiene texto: no es fila de datos
        v = [float(t.replace(",", "")) for t in toks]
        if len(toks) == 6:
            code, st, ft, _m, lbs, _kg = v
            row = {"code": int(code), "stages": int(st),
                   "stages_compression": None, "stages_ar_floater": None,
                   "stages_ar_compression": None,
                   "length_ft": ft, "weight_lbs": lbs}
        elif len(toks) == 7:
            # bombas grandes de construcción solo compression: 'stages'
            # (PK) toma el valor compression estándar
            code, st_co, ar_co, ft, _m, lbs, _kg = v
            row = {"code": int(code), "stages": int(st_co),
                   "stages_compression": int(st_co),
                   "stages_ar_floater": None,
                   "stages_ar_compression": int(ar_co),
                   "length_ft": ft, "weight_lbs": lbs}
        else:
            code, st_fl, st_co, ar_fl, ar_co, ft, _m, lbs, _kg = v
            row = {"code": int(code), "stages": int(st_fl),
                   "stages_compression": int(st_co),
                   "stages_ar_floater": int(ar_fl),
                   "stages_ar_compression": int(ar_co),
                   "length_ft": ft, "weight_lbs": lbs}
        if row["code"] > 500:  # descarta líneas numéricas que no son housings
            continue
        rows.append(row)
    return rows


def parse_pumps(text: str) -> list[dict]:
    """Divide el texto en fichas (una por modelo) y parsea cada una."""
    matches = list(MODEL_RE.finditer(text))
    pumps = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end]
        # una ficha real siempre tiene el rango óptimo; el índice no
        if "Optimum operating range" not in chunk:
            continue
        p: dict = {"model": m.group(1)}
        sm = SERIES_RE.search(chunk)
        p["series"] = sm.group(1) if sm else None
        for key, pat in FIELD_PATTERNS.items():
            mm = re.search(pat, chunk)
            if not mm:
                p[key] = None
            elif key in ("opt_range", "floater_range"):
                p[key] = (_num(mm.group(1)), _num(mm.group(2)))
            elif key == "stage_type":
                p[key] = mm.group(1)
            else:
                p[key] = _num(mm.group(1))
        p["housings"] = parse_housing_rows(chunk)
        pumps.append(p)
    return pumps


# ----------------------------------------------------------------------
# 2. Escritura en la base Excel (idempotente)
# ----------------------------------------------------------------------
def sheet_headers(ws) -> list[str]:
    return [c.value for c in ws[1] if c.value]


def ensure_columns(ws, new_cols: list[str]) -> dict[str, int]:
    """Agrega columnas si faltan; devuelve {nombre: índice 1-based}."""
    headers = sheet_headers(ws)
    for col in new_cols:
        if col not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            # copia el estilo del encabezado vecino
            ref = ws.cell(row=1, column=1)
            c = ws.cell(row=1, column=len(headers) + 1)
            c.font = ref.font.copy()
            c.fill = ref.fill.copy()
            headers.append(col)
    return {h: i + 1 for i, h in enumerate(headers)}


def delete_rows_where(ws, col_idx: int, value: str) -> int:
    """Borra las filas cuyo valor en col_idx == value (idempotencia)."""
    to_delete = [r for r in range(2, ws.max_row + 1)
                 if str(ws.cell(row=r, column=col_idx).value) == value]
    for r in reversed(to_delete):
        ws.delete_rows(r)
    return len(to_delete)


def append_row(ws, colmap: dict[str, int], data: dict) -> None:
    r = ws.max_row + 1
    for col, idx in colmap.items():
        v = data.get(col)
        cell = ws.cell(row=r, column=idx, value=v)
        cell.font = CELL_FONT


def main() -> None:
    pdf = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PDF
    if not pdf.exists():
        sys.exit(f"No se encuentra el PDF: {pdf} "
                 f"(pasá la ruta como argumento)")
    pumps = parse_pumps(pdf_text(pdf))
    if not pumps:
        sys.exit("No se extrajo ninguna bomba: revisar el parser")

    # --- manufacturers -------------------------------------------------
    wb = load_workbook(DATA_DIR / "manufacturers.xlsx")
    ws = wb["manufacturers"]
    cm = ensure_columns(ws, [])
    delete_rows_where(ws, cm["manufacturer_id"], MANUFACTURER)
    append_row(ws, cm, {
        "manufacturer_id": MANUFACTURER,
        "full_name": "Alkhorayef Petroleum Company",
        "country": "Arabia Saudita",
        "notes": "Importado del Product Catalog 2019 (línea SPECTRUM)"})
    wb.save(DATA_DIR / "manufacturers.xlsx")

    # --- data_sources ---------------------------------------------------
    wb = load_workbook(DATA_DIR / "data_sources.xlsx")
    ws = wb["data_sources"]
    cm = ensure_columns(ws, [])
    delete_rows_where(ws, cm["source_id"], SOURCE_ID)
    append_row(ws, cm, {"source_id": SOURCE_ID, "type": "catálogo",
                        "citation": CITATION})
    wb.save(DATA_DIR / "data_sources.xlsx")

    # --- equipment_series -------------------------------------------------
    wb = load_workbook(DATA_DIR / "equipment_series.xlsx")
    ws = wb["equipment_series"]
    cm = ensure_columns(ws, [])
    existing = {str(ws.cell(row=r, column=cm["series_id"]).value)
                for r in range(2, ws.max_row + 1)}
    added_series = []
    for serie in sorted({p["series"] for p in pumps if p["series"]}):
        if serie not in existing:
            append_row(ws, cm, {
                "series_id": serie,
                "od_nominal_inches": round(int(serie) / 100.0, 2),
                "used_in": "pumps",
                "notes": f"Agregada por importación {MANUFACTURER}"})
            added_series.append(serie)
    wb.save(DATA_DIR / "equipment_series.xlsx")

    # --- pumps + pump_housings ----------------------------------------------
    wb = load_workbook(DATA_DIR / "pumps.xlsx")
    ws_p, ws_h = wb["pumps"], wb["pump_housings"]
    cm_p = ensure_columns(ws_p, NEW_PUMP_COLS)
    cm_h = ensure_columns(ws_h, NEW_HOUSING_COLS)
    # idempotencia: fuera las filas Alkhorayef previas
    delete_rows_where(ws_p, cm_p["manufacturer_id"], MANUFACTURER)
    pref = MANUFACTURER + "_"
    rows_h = [r for r in range(2, ws_h.max_row + 1)
              if str(ws_h.cell(row=r, column=cm_h["pump_id"]).value
                     ).startswith(pref)]
    for r in reversed(rows_h):
        ws_h.delete_rows(r)

    skipped = []
    for p in pumps:
        if not p["series"] or not p["opt_range"] or not p["bep"]:
            skipped.append(p["model"])
            continue
        pid = f"{MANUFACTURER}_{p['series']}_{p['model']}"
        stages_list = [h["stages"] for h in p["housings"]]
        append_row(ws_p, cm_p, {
            "pump_id": pid, "manufacturer_id": MANUFACTURER,
            "series": p["series"], "model": p["model"],
            "od_inches": p["housing_od"],
            "min_flow_bpd": p["opt_range"][0],
            "max_flow_bpd": p["opt_range"][1],
            "bep_flow_bpd": p["bep"],
            "max_stages": max(stages_list) if stages_list else None,
            "source_id": SOURCE_ID,
            "min_flow_floater_bpd": p["floater_range"][0] if p["floater_range"] else None,
            "max_flow_floater_bpd": p["floater_range"][1] if p["floater_range"] else None,
            "min_casing_size_in": p["min_casing"],
            "housing_pressure_limit_psi": p["p_limit"],
            "housing_pressure_limit_high_psi": p["p_limit_high"],
            "shaft_diameter_in": p["shaft_d"],
            "shaft_hp_limit_std": p["shaft_std"],
            "shaft_hp_limit_hs": p["shaft_hs"],
            "shaft_hp_limit_uhs": p["shaft_uhs"],
            "stage_type": p["stage_type"],
        })
        seen_stages = set()
        for h in p["housings"]:
            if h["stages"] in seen_stages:
                continue  # PK (pump_id, stages)
            seen_stages.add(h["stages"])
            append_row(ws_h, cm_h, {
                "pump_id": pid, "stages": h["stages"],
                "housing_code": h["code"],
                "stages_compression": h["stages_compression"],
                "stages_ar_floater": h["stages_ar_floater"],
                "stages_ar_compression": h["stages_ar_compression"],
                "length_ft": h["length_ft"],
                "weight_lbs": h["weight_lbs"]})
    wb.save(DATA_DIR / "pumps.xlsx")

    n = len(pumps) - len(skipped)
    print(f"IMPORTADO {MANUFACTURER}: {n} bombas "
          f"({len(skipped)} salteadas: {skipped or '—'}), "
          f"{len(added_series)} series nuevas: {added_series}")
    print("Curvas: pendientes de digitalización (gráficos en el PDF) — "
          "las bombas quedan en la base pero no disponibles para diseño.")


if __name__ == "__main__":
    main()
