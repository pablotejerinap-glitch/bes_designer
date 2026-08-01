"""
Loader de catálogos desde Excel (.xlsx)
=======================================

Provee ``ExcelCatalogManager``: la misma interfaz que
``catalogs.loader.CatalogManager`` pero leyendo los datos desde los
archivos Excel de ``database_migration/data_excel/`` en vez de JSON.

CÓMO FUNCIONA (para la defensa de tesis)
----------------------------------------
``ExcelCatalogManager`` HEREDA de ``CatalogManager``. Solo reemplaza el
constructor (``__init__``): en lugar de leer JSON, lee Excel y arma
exactamente las mismas estructuras internas (``self._pumps``,
``self._motors``, etc.). Todos los métodos de consulta y selección de
equipos (get_motor, get_cable, get_seal...) se heredan SIN CAMBIOS.

Esto se llama *separación entre acceso a datos y lógica de negocio*:
la lógica de selección de equipos no sabe ni le importa de dónde
vienen los datos. Mañana un ``SqliteCatalogManager`` puede hacer lo
mismo contra una base SQLite sin tocar una línea de la lógica.

USO
---
    from database_migration.excel_loader import ExcelCatalogManager
    catalog = ExcelCatalogManager()          # lee data_excel/
    pumps = catalog.get_all_pumps()          # misma API de siempre
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from bes.catalogs.loader import CatalogManager
from bes.core.models import PumpCurve, PumpPerformancePoint

_EXCEL_DIR = Path(__file__).parent / "data_excel"


# ----------------------------------------------------------------------
# Utilidades de lectura
# ----------------------------------------------------------------------
def _read_sheet(path: Path, sheet: str) -> list[dict]:
    """Lee una hoja y devuelve una lista de dicts {columna: valor}.

    La fila 1 son los encabezados; las celdas vacías quedan como None
    (igual que ``null`` en JSON).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    out = []
    for row in rows:
        if all(v is None for v in row):
            continue  # ignora filas totalmente vacías
        out.append(dict(zip(headers, row)))
    wb.close()
    return out


def _split_ints(text) -> list[int]:
    """'50, 75, 100' -> [50, 75, 100]. Tolera celdas numéricas sueltas."""
    if text is None:
        return []
    if isinstance(text, (int, float)):
        return [int(text)]
    return [int(x.strip()) for x in str(text).split(",") if x.strip()]


def _split_strs(text) -> list[str]:
    """'456, 540' -> ['456', '540'] (series de motor como texto)."""
    if text is None:
        return []
    if isinstance(text, (int, float)):
        return [str(int(text))]
    return [x.strip() for x in str(text).split(",") if x.strip()]


def _rename_source(entry: dict) -> dict:
    """Devuelve el dict con 'source'→'_source' (convención del código).

    En Excel la columna se llama ``source`` (sin guión bajo, más
    amigable); internamente el código usa ``_source``.
    """
    out = {}
    for k, v in entry.items():
        if k in ("source", "range_source"):
            out["_" + k] = v
        else:
            out[k] = v
    return out


# ----------------------------------------------------------------------
# Manager
# ----------------------------------------------------------------------
class ExcelCatalogManager(CatalogManager):
    """Igual que CatalogManager, pero los datos vienen de Excel.

    Args:
        excel_dir: Carpeta con los .xlsx. Por defecto
            ``database_migration/data_excel/``.
    """

    def __init__(self, excel_dir: Optional[str] = None) -> None:
        base = Path(excel_dir) if excel_dir else _EXCEL_DIR
        # NOTA: no se llama a super().__init__() a propósito: el
        # constructor de la clase base leería los JSON. Acá armamos
        # las mismas estructuras internas desde Excel.
        self._pumps = self._load_pumps(base / "pumps.xlsx")
        self._motors = self._load_flat(base / "motors.xlsx", "motors")
        self._cables = self._load_cables(base / "cables.xlsx")
        self._seals = self._load_seals(base / "seals.xlsx")
        self._gas_handlers = self._load_flat(
            base / "gas_handlers.xlsx", "gas_handlers", optional=True
        )
        self._sensors = self._load_flat(
            base / "sensors.xlsx", "sensors", optional=True
        )

    # -- bombas: hoja maestra + hoja de curvas -------------------------
    @staticmethod
    def _load_pumps(path: Path) -> list[PumpCurve]:
        masters = _read_sheet(path, "pumps")
        curve_rows = _read_sheet(path, "performance_curves")

        # agrupa los puntos de curva por pump_id, respetando el orden
        curves: dict[str, list[PumpPerformancePoint]] = {}
        for r in curve_rows:
            curves.setdefault(r["pump_id"], []).append(
                PumpPerformancePoint(
                    flow_rate=r["flow_bpd"],
                    head_per_stage=r["head_ft_per_stage"],
                    hp_per_stage=r["hp_per_stage"],
                    efficiency=r["efficiency"],
                )
            )

        pumps = []
        for m in masters:
            pid = m["pump_id"]
            if pid not in curves:
                raise ValueError(
                    f"La bomba '{pid}' no tiene puntos en la hoja "
                    f"'performance_curves' de {path.name}"
                )
            pumps.append(
                PumpCurve(
                    manufacturer=m["manufacturer"],
                    series=str(m["series"]),
                    model=str(m["model"]),
                    od=m["od_inches"],
                    min_flow=m["min_flow_bpd"],
                    max_flow=m["max_flow_bpd"],
                    bep_flow=m["bep_flow_bpd"],
                    max_stages=m["max_stages"],
                    housing_options=_split_ints(m["housing_options"]),
                    points=curves[pid],
                )
            )
        return pumps

    # -- cables: hoja maestra + hoja de caída de tensión ----------------
    @staticmethod
    def _load_cables(path: Path) -> list[dict]:
        masters = _read_sheet(path, "cables")
        vdrop_rows = _read_sheet(path, "voltage_drop")

        # reconstruye el dict {temp: v_drop} con claves de texto,
        # exactamente como estaba en el JSON original
        vdrops: dict[str, dict[str, float]] = {}
        for r in vdrop_rows:
            vdrops.setdefault(r["cable_id"], {})[str(int(r["temp_f"]))] = (
                r["v_per_amp_per_1000ft"]
            )

        cables = []
        for m in masters:
            cid = m.pop("cable_id")
            if cid not in vdrops:
                raise ValueError(
                    f"El cable '{cid}' no tiene datos en la hoja "
                    f"'voltage_drop' de {path.name}"
                )
            c = _rename_source(m)
            c["size"] = str(c["size"])
            c["voltage_drop_v_per_amp_per_1000ft"] = vdrops[cid]
            cables.append(c)
        return cables

    # -- sellos: lista de series compatibles como texto -----------------
    @staticmethod
    def _load_seals(path: Path) -> list[dict]:
        seals = []
        for m in _read_sheet(path, "seals"):
            s = _rename_source(m)
            s["series"] = str(s["series"])
            s["compatible_motor_series"] = _split_strs(
                s["compatible_motor_series"]
            )
            seals.append(s)
        return seals

    # -- catálogos planos ------------------------------------------------
    @staticmethod
    def _load_flat(path: Path, sheet: str, optional: bool = False) -> list[dict]:
        if optional and not path.exists():
            return []
        out = []
        for m in _read_sheet(path, sheet):
            e = _rename_source(m)
            if "series" in e:
                e["series"] = str(e["series"])
            out.append(e)
        return out
