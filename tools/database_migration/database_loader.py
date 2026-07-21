"""
Loader de la base de datos Excel (esquema v2, normalizado)
==========================================================

Provee ``ExcelCatalogManager``: la misma interfaz que
``catalogs.loader.CatalogManager``, leyendo la base normalizada de
``data_excel/`` (ver DISENO_BASE_DE_DATOS.md).

Qué hace además de leer celdas (los "JOIN" del mundo Excel):
* pumps + pump_curves + pump_housings  → objetos PumpCurve completos
* cables + cable_voltage_drop          → dict de caída de tensión por cable
* seals + seal_motor_compatibility     → lista de series compatibles por sello
* manufacturer_id                      → campo 'manufacturer' del código

Toda la lógica de selección de equipos se HEREDA sin cambios de
CatalogManager: separación entre acceso a datos y lógica de ingeniería.

Supersede a ``excel_loader.py`` (esquema v1).

USO:
    from database_migration.database_loader import ExcelCatalogManager
    catalog = ExcelCatalogManager()
    pumps = catalog.get_all_pumps()   # misma API de siempre
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from bes.catalogs.loader import CatalogManager
from bes.core.models import PumpCurve, PumpPerformancePoint

_EXCEL_DIR = Path(__file__).parent / "data_excel"


def _read_sheet(path: Path, sheet: str) -> list[dict]:
    """Hoja → lista de dicts {columna: valor}; celda vacía → None."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]
    out = [dict(zip(headers, row)) for row in rows
           if not all(v is None for v in row)]
    wb.close()
    return out


def _to_entry(rec: dict, id_col: str) -> dict:
    """Convierte una fila Excel al formato de dict que espera el código.

    * elimina la PK legible (id_col) — el código no la usa,
    * manufacturer_id → manufacturer,
    * source/range_source → _source/_range_source (convención del código),
    * series numérica → texto (Excel guarda '400' como número).
    """
    out = {}
    for k, v in rec.items():
        if k == id_col:
            continue
        if k == "manufacturer_id":
            out["manufacturer"] = v
        elif k in ("source", "range_source"):
            out["_" + k] = v
        else:
            out[k] = v
    if "series" in out and out["series"] is not None:
        out["series"] = str(out["series"])
    return out


class ExcelCatalogManager(CatalogManager):
    """Misma interfaz que CatalogManager; datos desde data_excel/ (v2)."""

    def __init__(self, excel_dir: Optional[str] = None) -> None:
        base = Path(excel_dir) if excel_dir else _EXCEL_DIR
        # No se llama a super().__init__(): leería los JSON. Se arman
        # las mismas estructuras internas desde la base normalizada.
        self._pumps = self._load_pumps(base / "pumps.xlsx")
        self._motors = [_to_entry(r, "motor_id")
                        for r in _read_sheet(base / "motors.xlsx", "motors")]
        self._cables = self._load_cables(base / "cables.xlsx")
        self._seals = self._load_seals(base / "seals.xlsx")
        self._gas_handlers = [
            _to_entry(r, "gas_handler_id")
            for r in _read_sheet(base / "gas_handlers.xlsx", "gas_handlers")
        ] if (base / "gas_handlers.xlsx").exists() else []
        self._sensors = [
            _to_entry(r, "sensor_id")
            for r in _read_sheet(base / "sensors.xlsx", "sensors")
        ] if (base / "sensors.xlsx").exists() else []

    # ------------------------------------------------------------------
    @staticmethod
    def _load_pumps(path: Path) -> list[PumpCurve]:
        masters = _read_sheet(path, "pumps")
        curves: dict[str, list[PumpPerformancePoint]] = {}
        for r in _read_sheet(path, "pump_curves"):
            curves.setdefault(r["pump_id"], []).append(
                PumpPerformancePoint(
                    flow_rate=r["flow_bpd"],
                    head_per_stage=r["head_ft_per_stage"],
                    hp_per_stage=r["hp_per_stage"],
                    efficiency=r["efficiency"],
                )
            )
        housings: dict[str, list[int]] = {}
        for r in _read_sheet(path, "pump_housings"):
            housings.setdefault(r["pump_id"], []).append(int(r["stages"]))

        pumps = []
        for m in masters:
            pid = m["pump_id"]
            if pid not in curves:
                raise ValueError(
                    f"La bomba '{pid}' no tiene filas en 'pump_curves'")
            if pid not in housings:
                raise ValueError(
                    f"La bomba '{pid}' no tiene filas en 'pump_housings'")
            pumps.append(PumpCurve(
                manufacturer=m["manufacturer_id"],
                series=str(m["series"]),
                model=str(m["model"]),
                od=m["od_inches"],
                min_flow=m["min_flow_bpd"],
                max_flow=m["max_flow_bpd"],
                bep_flow=m["bep_flow_bpd"],
                max_stages=m["max_stages"],
                housing_options=housings[pid],
                points=curves[pid],
            ))
        return pumps

    @staticmethod
    def _load_cables(path: Path) -> list[dict]:
        masters = _read_sheet(path, "cables")
        vdrops: dict[str, dict[str, float]] = {}
        for r in _read_sheet(path, "cable_voltage_drop"):
            vdrops.setdefault(r["cable_id"], {})[str(int(r["temp_f"]))] = (
                r["v_per_amp_per_1000ft"])
        cables = []
        for m in masters:
            cid = m["cable_id"]
            if cid not in vdrops:
                raise ValueError(
                    f"El cable '{cid}' no tiene filas en 'cable_voltage_drop'")
            c = _to_entry(m, "cable_id")
            c["size"] = str(c["size"])
            c["voltage_drop_v_per_amp_per_1000ft"] = vdrops[cid]
            cables.append(c)
        return cables

    @staticmethod
    def _load_seals(path: Path) -> list[dict]:
        compat: dict[str, list[str]] = {}
        for r in _read_sheet(path, "seal_motor_compatibility"):
            compat.setdefault(r["seal_id"], []).append(str(r["motor_series"]))
        seals = []
        for m in _read_sheet(path, "seals"):
            sid = m["seal_id"]
            s = _to_entry(m, "seal_id")
            s["compatible_motor_series"] = compat.get(sid, [])
            seals.append(s)
        return seals

    # ------------------------------------------------------------------
    # Tablas nuevas del esquema v2 (aún no usadas por la aplicación)
    # ------------------------------------------------------------------
    def get_transformer_sizes_kva(
        self, excel_dir: Optional[str] = None
    ) -> list[float]:
        """Tamaños estándar [kVA] desde transformers.xlsx (reemplazará a
        la tupla hardcodeada de core/electrical.py al integrar)."""
        base = Path(excel_dir) if excel_dir else _EXCEL_DIR
        rows = _read_sheet(base / "transformers.xlsx", "transformers")
        return sorted(float(r["kva_rating"]) for r in rows)
