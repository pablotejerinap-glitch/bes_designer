"""Excel report exporter for BES/ESP design results using openpyxl."""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bes.core.models import DesignResult

_BLUE_FILL    = PatternFill("solid", fgColor="1a237e")
_LT_BLUE_FILL = PatternFill("solid", fgColor="e3f2fd")
_WHITE_FILL   = PatternFill("solid", fgColor="ffffff")
_HDR_FONT     = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_LABEL_FONT   = Font(bold=True, name="Calibri", size=10)
_DATA_FONT    = Font(name="Calibri", size=10)
_THIN         = Side(border_style="thin", color="BDBDBD")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _section_title(ws, row: int, title: str, span: int = 3) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = _BLUE_FILL
    c.font = _HDR_FONT
    c.alignment = _CENTER
    c.border = _BORDER


def _header_row(ws, row: int, headers: list) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _BLUE_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BORDER


def _data_row(ws, row: int, label: str, value, unit: str = "", alt: bool = False) -> None:
    fill = _LT_BLUE_FILL if alt else _WHITE_FILL
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _LABEL_FONT; lc.fill = fill; lc.alignment = _LEFT; lc.border = _BORDER
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = _DATA_FONT;  vc.fill = fill; vc.alignment = _LEFT; vc.border = _BORDER
    if unit:
        uc = ws.cell(row=row, column=3, value=unit)
        uc.font = _DATA_FONT; uc.fill = fill; uc.alignment = _LEFT; uc.border = _BORDER


def _set_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_resumen(ws, dr: DesignResult, well_data: dict) -> None:
    ws.title = "Resumen"
    _set_widths(ws, [34, 22, 12])

    objectives = well_data.get("objectives")
    reservoir  = well_data.get("reservoir")

    row = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    c = ws.cell(row=1, column=1, value="BES Designer — Reporte de Diseño")
    c.font = Font(bold=True, size=14, color="1a237e", name="Calibri")
    c.alignment = _CENTER
    row = 2
    ws.cell(row=row, column=1, value=f"Generado: {date.today()}").font = Font(
        size=9, color="757575", name="Calibri"
    )
    row = 4

    _section_title(ws, row, "RESUMEN EJECUTIVO", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    pairs = []
    if reservoir:
        pairs.append(("Presión reservorio", round(reservoir.static_pressure, 0), "psi"))
    if objectives:
        pairs.append(("Tasa objetivo", round(objectives.target_flow_rate, 0), "STB/d"))
    pairs += [
        ("Bomba seleccionada",  f"{dr.pump_manufacturer} {dr.pump_model}", "—"),
        ("Número de etapas",    dr.num_stages,                              "—"),
        ("HP del motor",        round(dr.motor_hp, 0),                      "hp"),
        ("Tasa alcanzada",      round(dr.flow_rate_achieved, 0),            "STB/d"),
        ("TDH",                 round(dr.total_head_required, 1),           "ft"),
        ("Eficiencia bomba",    f"{dr.pump_efficiency:.1%}",                "—"),
        ("Eficiencia sistema",  f"{dr.system_efficiency:.1%}",              "—"),
        ("Transformador",       round(dr.transformer_kva, 0),               "kVA"),
        ("Voltaje motor",       round(dr.motor_voltage, 0),                 "V"),
        ("GIP en intake",       f"{dr.gip_fraction:.1%}",                   "fracción"),
    ]
    for i, (label, value, unit) in enumerate(pairs):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1

    if dr.warnings:
        row += 1
        _section_title(ws, row, "ADVERTENCIAS", 3)
        row += 1
        for w in dr.warnings:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c = ws.cell(row=row, column=1, value=f"• {w}")
            c.font = Font(color="E65100", name="Calibri", size=10)
            c.alignment = _LEFT
            row += 1

    # Crédito del autor al pie, igual que el pie de página del PDF.
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(
        row=row, column=1,
        value="BES Designer v2.0 — Desarrollado por Pablo Agustín Tejerina, UNCo",
    )
    c.font = Font(size=8, color="757575", name="Calibri", italic=True)
    c.alignment = _LEFT


def _build_inputs(ws, well_data: dict) -> None:
    ws.title = "Inputs"
    _set_widths(ws, [34, 22, 12])

    reservoir  = well_data.get("reservoir")
    fluid      = well_data.get("fluid")
    well       = well_data.get("well")
    surface    = well_data.get("surface")
    objectives = well_data.get("objectives")
    row = 1

    def section(title, pairs):
        nonlocal row
        _section_title(ws, row, title, 3)
        row += 1
        _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
        row += 1
        for i, (label, value, unit) in enumerate(pairs):
            _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
            row += 1
        row += 1

    if reservoir:
        section("RESERVORIO", [
            ("Presión estática",    round(reservoir.static_pressure, 0),    "psi"),
            ("Presión de burbuja",  round(reservoir.bubble_point, 0),        "psi"),
            ("PI",                  round(reservoir.productivity_index, 3),  "STB/d/psi"),
            ("Método IPR",          reservoir.ipr_method.name,               "—"),
            ("Temperatura",         round(reservoir.reservoir_temp, 0),      "°F"),
            ("Mecanismo de empuje", reservoir.drive_mechanism.name,          "—"),
        ])

    if fluid:
        section("FLUIDO", [
            ("API del crudo",       round(fluid.oil_api, 1),                          "°API"),
            ("Corte de agua (WC)", f"{fluid.water_cut:.1%}",                          "—"),
            ("GOR",                 round(fluid.gor, 0),                               "scf/STB"),
            ("Gr. esp. gas",        round(fluid.gas_sg, 3),                           "—"),
            ("Gr. esp. agua",       round(fluid.water_sg, 3),                         "—"),
            ("Viscosidad dead oil", round(fluid.oil_viscosity_dead, 2),               "cp"),
            ("H2S",                 round(fluid.h2s_content, 0),                      "ppm"),
            ("CO2",                 round(fluid.co2_content, 0),                      "ppm"),
            ("Producción de arena", "Sí" if fluid.sand_production else "No",          "—"),
        ])

    if well:
        section("GEOMETRÍA DEL POZO", [
            ("Profundidad total",    round(well.total_depth, 0),            "ft"),
            ("OD casing",           round(well.casing_od, 3),              "in"),
            ("Peso casing",         round(well.casing_weight, 1),          "lb/ft"),
            ("ID casing",           round(well.casing_id, 3),              "in"),
            ("OD tubing",           round(well.tubing_od, 3),              "in"),
            ("ID tubing",           round(well.tubing_id, 3),              "in"),
            ("Perforaciones techo", round(well.perforations_top, 0),       "ft"),
            ("Perforaciones base",  round(well.perforations_bottom, 0),    "ft"),
            ("Desviación máx.",     round(well.deviation_max, 1),          "°"),
            ("Temp. cabezal",       round(well.wellhead_temp, 0),          "°F"),
            ("BHT",                 round(reservoir.reservoir_temp, 0),    "°F"),
        ])

    if surface:
        section("CONDICIONES SUPERFICIALES", [
            ("Presión cabezal req.", round(surface.wellhead_pressure_required, 0), "psi"),
            ("Longitud flowline",    round(surface.flowline_length, 0),            "ft"),
            ("ID flowline",          round(surface.flowline_id, 2),                "in"),
            ("Elevación flowline",   round(surface.flowline_elevation_change, 0),  "ft"),
            ("Presión separador",    round(surface.separator_pressure, 0),         "psi"),
            ("Voltaje disponible",   round(surface.power_supply_voltage, 0),       "V"),
            ("Frecuencia",           round(surface.frequency, 0),                  "Hz"),
        ])

    if objectives:
        section("OBJETIVOS DE DISEÑO", [
            ("Tasa objetivo",       round(objectives.target_flow_rate, 0),          "STB/d"),
            ("Margen seguridad",    round(objectives.safety_margin_depth, 0),       "ft"),
            ("Venteo de gas",       "Sí" if objectives.allow_gas_venting else "No", "—"),
            ("Máx. GIP",           f"{objectives.max_gip:.1%}",                    "—"),
            ("Vida útil",           round(objectives.design_life_years, 1),         "años"),
            ("VSD",                 "Sí" if objectives.use_vsd else "No",           "—"),
        ])


def _build_tdh(ws, dr: DesignResult, well_data: dict) -> None:
    ws.title = "TDH"
    _set_widths(ws, [34, 22, 12])

    reservoir  = well_data.get("reservoir")
    fluid      = well_data.get("fluid")
    well       = well_data.get("well")
    surface    = well_data.get("surface")
    objectives = well_data.get("objectives")

    row = 1
    _section_title(ws, row, "DESGLOSE DEL TDH", 3)
    row += 1
    _header_row(ws, row, ["Componente", "Valor", "Unidad"])
    row += 1

    if all(x is not None for x in (reservoir, fluid, well, surface, objectives)):
        try:
            from bes.core.tdh import calculate_tdh
            tdh = calculate_tdh(
                reservoir=reservoir, fluid=fluid, well=well,
                surface=surface, objectives=objectives,
                pump_depth=dr.pump_setting_depth, pip=dr.intake_pressure,
            )
            comps = [
                ("Elevación vertical",     round(tdh["vertical_lift_ft"], 1),           "ft"),
                ("Fricción en tubería",    round(tdh["tubing_friction_ft"], 1),          "ft"),
                ("Presión cabezal (head)", round(tdh["wellhead_pressure_head_ft"], 1),   "ft"),
                ("PIP (head)",             round(tdh["pip_head_ft"], 1),                 "ft"),
            ]
            for i, (label, value, unit) in enumerate(comps):
                _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
                row += 1

            # Total row — bold blue
            for col, val in enumerate([("TOTAL TDH", round(tdh["tdh_ft"], 1), "ft")][0], start=1):
                pass
            for col, val in enumerate(["TOTAL TDH", round(tdh["tdh_ft"], 1), "ft"], start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(bold=True, name="Calibri", size=10)
                c.fill = _LT_BLUE_FILL
                c.alignment = _LEFT
                c.border = _BORDER
            row += 2

            _section_title(ws, row, "PARÁMETROS ADICIONALES", 3)
            row += 1
            _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
            row += 1
            for i, (label, value, unit) in enumerate([
                ("SG líquido",        round(tdh["sg_liquid"], 3),     "—"),
                ("Profundidad bomba", round(tdh["pump_depth_ft"], 0), "ft"),
                ("PIP",               round(tdh["pip_psi"], 0),       "psi"),
            ]):
                _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
                row += 1
            return
        except Exception:
            pass

    _data_row(ws, row, "TDH (diseño)", round(dr.total_head_required, 1), "ft")
    row += 1
    _data_row(ws, row, "PIP", round(dr.intake_pressure, 0), "psi")


def _build_bomba(ws, dr: DesignResult) -> None:
    ws.title = "Bomba"
    _set_widths(ws, [34, 22, 12])

    row = 1
    _section_title(ws, row, "BOMBA SELECCIONADA", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    for i, (label, value, unit) in enumerate([
        ("Fabricante",              dr.pump_manufacturer,             "—"),
        ("Serie",                   dr.pump_series,                   "—"),
        ("Modelo",                  dr.pump_model,                    "—"),
        ("OD",                      round(dr.pump_od, 3),             "in"),
        ("Número de etapas",        dr.num_stages,                    "—"),
        ("Carcasas (housings)",     " + ".join(
            f"{h['stages']} et" + (f" [{h['code']}]" if h["code"] else "")
            for h in dr.housing_detail
        ) or "—",                                                     "—"),
        ("Capacidad instalada",     dr.housing_size_stages,           "etapas"),
        ("Etapas ciegas (dummy)",   dr.dummy_stages,                  "etapas"),
        ("MaxP carcasa superior",   round(dr.max_housing_pressure_psi, 0), "psi"),
        ("Presión admisible",       round(dr.housing_pressure_limit_psi, 0)
                                    if dr.housing_pressure_verified else "sin dato", "psi"),
        ("Verificación de presión", ("OK" if dr.housing_pressure_ok else "FAIL")
                                    if dr.housing_pressure_verified else "sin verificar", "—"),
        ("Criterio de carcasas",    dr.housing_rationale or "—",      "—"),
        ("Eficiencia bomba",        f"{dr.pump_efficiency:.1%}",      "—"),
        ("Head / etapa",            round(dr.head_per_stage, 1),      "ft/etapa"),
        ("HP / etapa",              round(dr.hp_per_stage, 3),        "hp/etapa"),
        ("HP total bomba",          round(dr.total_pump_hp, 1),       "hp"),
        ("Profundidad instalación", round(dr.pump_setting_depth, 0),  "ft"),
        ("PIP (intake)",            round(dr.intake_pressure, 0),     "psi"),
        ("GIP en intake",           f"{dr.gip_fraction:.1%}",         "fracción"),
        ("Tasa alcanzada",          round(dr.flow_rate_achieved, 0),  "STB/d"),
        ("Frecuencia operación",    round(dr.operating_frequency, 0), "Hz"),
    ]):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1


def _build_electrico(ws, dr: DesignResult) -> None:
    ws.title = "Electrico"
    _set_widths(ws, [34, 22, 12])

    row = 1
    _section_title(ws, row, "MOTOR", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    for i, (label, value, unit) in enumerate([
        ("Fabricante", dr.motor_manufacturer,       "—"),
        ("Modelo",     dr.motor_model,              "—"),
        ("HP nominal", round(dr.motor_hp, 0),       "hp"),
        ("Voltaje",    round(dr.motor_voltage, 0),  "V"),
        ("Amperaje",   round(dr.motor_amperage, 0), "A"),
        ("OD motor",   round(dr.motor_od, 3),       "in"),
        ("Longitud",   round(dr.motor_length, 1),   "ft"),
    ]):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1

    row += 1
    _section_title(ws, row, "SELLO / PROTECTOR", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    if dr.seal_model:
        seal_rows = [
            ("Fabricante",             dr.seal_manufacturer,                 "—"),
            ("Modelo",                 dr.seal_model,                        "—"),
            ("Tipo",                   dr.seal_type,                         "—"),
            ("Capacidad de empuje",    round(dr.seal_thrust_capacity_lbs, 0), "lbs"),
            ("Empuje axial estimado",  round(dr.axial_thrust_lbs, 0),        "lbs"),
        ]
    else:
        seal_rows = [("Protector", "Sin opción compatible en catálogo", "—")]
    for i, (label, value, unit) in enumerate(seal_rows):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1

    row += 1
    _section_title(ws, row, "CABLE ELÉCTRICO", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    for i, (label, value, unit) in enumerate([
        ("Tipo de cable",            dr.cable_type,                         "—"),
        ("Calibre (AWG)",           f"#{dr.cable_awg}",                    "—"),
        ("Caída de tensión",         round(dr.cable_voltage_drop, 0),       "V"),
        ("Voltaje superficial req.", round(dr.surface_voltage_required, 0), "V"),
    ]):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1

    row += 1
    _section_title(ws, row, "TRANSFORMADOR Y SISTEMA", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    for i, (label, value, unit) in enumerate([
        ("Transformador",          round(dr.transformer_kva, 0),       "kVA"),
        ("Eficiencia del sistema", f"{dr.system_efficiency:.1%}",      "—"),
    ]):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1

    row += 1
    _section_title(ws, row, "MANEJO DE GAS Y MONITOREO", 3)
    row += 1
    _header_row(ws, row, ["Parámetro", "Valor", "Unidad"])
    row += 1

    gh_val = (
        f"{dr.gas_handler_manufacturer} {dr.gas_handler_model} "
        f"({dr.gas_handler_type}, η {dr.gas_handler_efficiency:.0%})"
        if dr.gas_handler_model else "No requerido (GIP bajo)"
    )
    sn_val = (
        f"{dr.sensor_manufacturer} {dr.sensor_model}"
        if dr.sensor_model else "No disponible en catálogo"
    )
    for i, (label, value, unit) in enumerate([
        ("Separador de gas", gh_val, "—"),
        ("Sensor de fondo",  sn_val, "—"),
    ]):
        _data_row(ws, row, label, value, unit, alt=(i % 2 == 1))
        row += 1


def _build_warnings(ws, dr: DesignResult) -> None:
    ws.title = "Warnings"
    _set_widths(ws, [80])

    all_w = list(dr.warnings)
    if dr.gip_fraction > 0.30:
        all_w.append(
            f"CRÍTICO — Gas libre en intake: {dr.gip_fraction:.1%}. "
            "Se requiere separador de gas de fondo."
        )
    elif dr.gip_fraction > 0.10:
        all_w.append(
            f"AVISO — Gas libre en intake: {dr.gip_fraction:.1%}. "
            "Considerar separador de gas."
        )

    row = 1
    c = ws.cell(row=row, column=1, value="ADVERTENCIAS Y RECOMENDACIONES")
    c.fill = _BLUE_FILL; c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _BORDER
    row += 1

    if not all_w:
        ws.cell(row=row, column=1, value="Sin advertencias.").font = _DATA_FONT
    else:
        for w in all_w:
            c = ws.cell(row=row, column=1, value=f"• {w}")
            c.font = Font(color="E65100", name="Calibri", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = _BORDER
            ws.row_dimensions[row].height = 30
            row += 1


def generate_design_excel(
    design_result: DesignResult,
    well_data: dict,
    output_path: Optional[str] = None,
) -> bytes | str:
    """Generate an Excel design report.

    Args:
        design_result: DesignResult from generate_recommendations().
        well_data: Dict with keys: reservoir, fluid, well, surface, objectives.
        output_path: Write to this path and return it. If None, return Excel bytes.
    """
    wb = Workbook()

    _build_resumen(wb.active, design_result, well_data)
    _build_inputs(wb.create_sheet(), well_data)
    _build_tdh(wb.create_sheet(), design_result, well_data)
    _build_bomba(wb.create_sheet(), design_result)
    _build_electrico(wb.create_sheet(), design_result)
    _build_warnings(wb.create_sheet(), design_result)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    if output_path is not None:
        with open(output_path, "wb") as fh:
            fh.write(xlsx_bytes)
        return output_path
    return xlsx_bytes
